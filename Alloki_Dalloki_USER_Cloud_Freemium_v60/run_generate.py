#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Alloki & Dalloki Generator (v31)
- v4 기반 생성기 + 개선:
✅ --thumb_pick A|B|C 로 썸네일 1종만 생성 가능
✅ --platform/--utm_campaign/--mode/free-lock 로 기존 동작 유지
✅ --story_preset / --story_last_preset 유지 (v4 기능 포함)
✅ Cards 엑셀: day,text,color,mood,price,cta 읽어서 반영
"""

from __future__ import annotations
import argparse, os, sys, zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from PIL import Image, ImageDraw, ImageFont, ImageFilter
import qrcode
import openpyxl


api_key = os.environ.get("OPENAI_API_KEY", "").strip()

OUT_SQUARE = (1080, 1080)
OUT_STORY  = (1080, 1920)
API_SIZE   = "1024x1024"
MODEL      = "gpt-image-1"


BASE_PROMPT = (
    "Two adorable pastel rainbow baby poodles, Alloki and Dalloki, "
    "sitting calmly side by side, gentle expressions, minimal background, "
    "ivory tone, clean composition, emotional but quiet mood, storybook style, high resolution. "
    "Leave generous empty space for text overlay. "
    "No text, no letters, no watermark."
)
SEASON_ADDONS = {
    "spring": "Soft peach and cream background, spring light.",
    "summer": "Soft mint and ivory background, cool calm mood.",
    "autumn": "Oatmeal and warm brown background, reflective mood.",
    "winter": "Ivory and light gray-blue background, soft winter light.",
    "yearend_bundle": "Four-season subtle gradient ring, premium calm feeling.",
}
THUMB_COPY_DEFAULT = {
    "A": "오늘의 마음을 꺼내보세요",
    "B": "지금 안 보면 놓쳐요",
    "C": "사계절을 건너온 마음",
}


def offer_plan(offer_code: str, season: str, days_arg: int, bonus_arg: int) -> Tuple[int,int,str]:
    """
    Returns (days, bonus, label)
    - D7/D14/D21 -> days fixed, bonus=0
    - SEASONPACK -> days=21, bonus default 3 (unless --bonus set)
    """
    oc = (offer_code or "").upper()
    if oc == "D7": return 7, 0, "7일"
    if oc == "D14": return 14, 0, "14일"
    if oc == "D21": return 21, 0, "21일"
    if oc == "SEASONPACK":
        b = bonus_arg if bonus_arg else 3
        return 21, b, f"{season} 시즌팩"
    return days_arg, bonus_arg, "custom"

def thumb_copy_for_offer(offer_code: str, season: str) -> Dict[str, str]:
    oc = (offer_code or "").upper()

    if oc == "SEASONPACK":
        season_kr = {
            "spring": "봄",
            "summer": "여름",
            "autumn": "가을",
            "winter": "겨울",
        }.get(season, season)


        return {
            "A": f"{season_kr} 시즌팩 21+3 오늘의 마음을 꺼내요",
            "B": f"{season_kr} 시즌팩 21+3 지금 안 사면 늦겠어요",
            "C": f"{season_kr} 시즌팩 21+3 프리미엄 한정",
        }

    if oc == "D7":
        return {"A": "7일 카드 · 오늘의 마음", "B": "7일 카드 · 지금 시작", "C": "7일 카드 · 가볍게"}

    if oc == "D14":
        return {"A": "14일 카드 · 마음 회복", "B": "14일 카드 · 놓치면 후회", "C": "14일 카드 · 프리미엄"}

    if oc == "D21":
        return {"A": "21일 카드 · 마음 루틴", "B": "21일 카드 · 지금이 타이밍", "C": "21일 카드 · 프리미엄"}

    return THUMB_COPY_DEFAULT.copy()


def ensure_dir(p: Path): p.mkdir(parents=True, exist_ok=True)

def compute_countdown(deadline: str, fallback_days: int) -> int:
    """
    deadline: YYYY-MM-DD
    returns D-N (>=0). If invalid/missing, return fallback_days.
    """
    if not deadline:
        return int(fallback_days)
    try:
        d = datetime.strptime(deadline, "%Y-%m-%d").date()
        today = date.today()
        delta = (d - today).days
        return max(0, int(delta))
    except Exception:
        return int(fallback_days)

def seasonpack_stage_labels(days_left: int) -> tuple[str,str]:
    """
    returns (urgency_tag, headline_suffix)
    """
    if days_left <= 0:
        return "오늘 마감", "마지막 기회"
    if days_left <= 1:
        return "D-1", "내일 마감"
    if days_left <= 3:
        return "D-3", "곧 마감"
    if days_left <= 7:
        return "D-7", "이번 주 마감"
    return f"D-{days_left}", ""

def seasonpack_cta_body_by_stage(days_left: int, segment: str) -> str:
    seg = (segment or "new").lower()
    if days_left <= 0:
        return "오늘 마감!\n지금 구매하면 보너스 3장 즉시 제공"
    if days_left <= 1:
        return "내일 마감!\n21일 + 보너스 3장 지금 받기"
    if days_left <= 3:
        return ("재구매자 전용 업그레이드!\n" if seg=="repeat" else "") + "21일 + 보너스 3장, 곧 마감"
    if days_left <= 7:
        return "이번 주 마감!\n3주 루틴 + 보너스 3장"
    return "3주 루틴 완성!\n보너스 3장까지 바로 받기"

def seasonpack_cta_t1_teaser_by_stage(days_left: int, platform: str, segment: str) -> str:
    seg = (segment or "new").lower()
    if days_left <= 0:
        return "오늘 마감!\n지금 열어보면 바로 사고 싶어져요"
    if days_left <= 1:
        return "내일 마감!\n놓치면 다음 시즌까지 기다려요"
    if days_left <= 3:
        return ("재구매자 전용" if seg=="repeat" else "처음이면 지금") + "\n곧 마감이라 가장 많이 사요"
    if days_left <= 7:
        return "이번 주 마감\n21일 + 보너스 3장 공개"
    # default
    return cta_t1_teaser_copy(platform, "", segment)

def add_countdown_label(im_path, out_path, days_left: int = 3):
    """Overlay D-N countdown near top-left."""
    from PIL import Image, ImageDraw, ImageFont
    im = Image.open(im_path).convert("RGBA")
    w,h = im.size
    draw = ImageDraw.Draw(im)
    label = f"D-{int(days_left)}"
    try:
        font = ImageFont.truetype(DEFAULT_FONT, int(h*0.07))
    except:
        font = ImageFont.load_default()
    pad = int(h*0.025)
    # background pill
    tw, th = draw.textsize(label, font=font)
    x0, y0 = pad, pad
    draw.rounded_rectangle([x0-pad, y0-pad, x0+tw+pad, y0+th+pad], radius=int(pad*1.2), fill=(0,0,0,170))
    draw.text((x0,y0), label, fill=(255,255,255,255), font=font)
    im.save(out_path, "PNG")

def cta_t1_teaser_copy(platform: str, season: str, segment: str) -> str:
    """Return teaser body for CTA_T1. Segment-based split."""
    seg = (segment or "new").lower()
    if seg == "repeat":
        return "재구매자 전용 루틴 업그레이드!\n21일 + 보너스 3장 공개"
    # new
    if (platform or "").lower() == "tiktok":
        return "처음이면 지금이 기회!\n21일 + 보너스 3장 공개"
    return "처음이라면 가볍게 시작!\n21일 + 보너스 3장 공개"

def make_cta_t1_mp4(
    story_png_path: Path,
    mp4_path: Path,
    seconds: float = 2.0,
    fps: int = 30,
    shake: bool=False,
    red_border: bool=False,
    shake_intensity: int = 3,
    banner_text: str = "",
    banner_color: tuple = (255,0,0),
    graph_bins: Optional[list] = None,
):
    """
    2s MP4: subtle zoom-in + sparkle dots, optional shake + red border + warning banner + mini graph.
    Requires imageio + imageio-ffmpeg + numpy.
    """
    import numpy as np
    import imageio.v3 as iio
    im0 = Image.open(story_png_path).convert("RGB")
    w,h = im0.size
    n = int(seconds * fps)
    frames = []
    border = 22 if red_border else 0

    # graph config
    bins = graph_bins or []
    g_w = int(w*0.34)
    g_h = int(h*0.16)
    g_x = int(w*0.06)
    g_y = int(h*0.74)

    for t in range(n):
        z = 1.0 + 0.06 * (t / max(1, n-1))
        zw, zh = int(w*z), int(h*z)
        imz = im0.resize((zw, zh), Image.BICUBIC)
        left = (zw - w)//2
        top = (zh - h)//2
        imc = imz.crop((left, top, left+w, top+h)).copy()

        if shake:
            amp = max(2, int(shake_intensity))
            dx = int(((-1)**t) * (amp + (t % 4)))
            dy = int(((-1)**(t+1)) * (amp + ((t+1) % 4)))
            tmp = Image.new("RGB",(w,h),(0,0,0))
            tmp.paste(imc, (dx,dy))
            imc = tmp

        draw = ImageDraw.Draw(imc)

        # sparkle dots
        for k in range(14):
            x = int((w*(k+1)/15) + (t*9 + k*31) % 27 - 13)
            y = int((h*(k+1)/15) + (t*13 + k*17) % 31 - 15)
            r = 2 + ((t + k) % 3)
            draw.ellipse([x-r,y-r,x+r,y+r], fill=(255,255,255))

        # warning banner
        if banner_text:
            banner_h = int(h*0.12)
            draw.rectangle([0,0,w,banner_h], fill=banner_color)
            f = pick_font(DEFAULT_FONT, int(banner_h*0.45))
            draw.text((int(w*0.05), int(banner_h*0.25)), banner_text, font=f, fill=(255,255,255))

        # mini purchase graph (simple polyline)
        if bins:
            draw.rounded_rectangle([g_x-10, g_y-10, g_x+g_w+10, g_y+g_h+10], radius=18, fill=(0,0,0,140))
            mx = max(1, max(bins))
            pts = []
            for i, v in enumerate(bins):
                px = g_x + int(i*(g_w/(max(1,len(bins)-1))))
                py = g_y + g_h - int((v/mx)*g_h)
                pts.append((px,py))
            if len(pts) >= 2:
                draw.line(pts, fill=(255,255,255), width=4)
            # label
            f2 = pick_font(DEFAULT_FONT, int(h*0.03))
            draw.text((g_x, g_y-34), "실시간 구매 그래프", font=f2, fill=(255,255,255))

        # border
        if red_border:
            draw.rectangle([0,0,w-1,h-1], outline=(255,0,0), width=max(10,border))

        frames.append(np.array(imc))

    mp4_path.parent.mkdir(parents=True, exist_ok=True)
    iio.imwrite(mp4_path, frames, fps=fps, codec="libx264", pixelformat="yuv420p")

def pick_font(path: Optional[str], size: int):
    if path and Path(path).exists():
        return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()

def header_index(ws):
    hdr = [c.value for c in ws[1]]
    idx = {}
    for i,v in enumerate(hdr):
        if v is None: continue
        idx[str(v).strip().lower()] = i
    return idx

def normalize_day(val) -> str:
    s = str(val).strip().upper()
    if s.startswith("DAY"):
        s = s[3:]
    s = "".join(ch for ch in s if ch.isdigit())
    return f"DAY{int(s):02d}" if s else ""

def load_cards_xlsx(path: Path, sheet: str) -> Dict[str, dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    idx = header_index(ws)
    if "day" not in idx or "text" not in idx:
        raise ValueError("Cards sheet must include headers: day, text (optional: color,mood,price,cta)")
    def get(row, key):
        j = idx.get(key)
        if j is None or j >= len(row): return ""
        v = row[j]
        return "" if v is None else str(v).strip()
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        day = normalize_day(row[idx["day"]])
        if not day: continue
        out[day] = {
            "text": get(row, "text"),
            "color": get(row, "color"),
            "mood": get(row, "mood"),
            "price": get(row, "price"),
            "cta": get(row, "cta"),
        }
    return out

def load_thumb_copy_xlsx(path: Path, sheet: str) -> Dict[str,str]:
    try:
        wb = openpyxl.load_workbook(path)
    except Exception:
        return {}
    if sheet not in wb.sheetnames:
        return {}
    ws = wb[sheet]
    idx = header_index(ws)
    if "variant" not in idx or "copy" not in idx:
        return {}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[idx["variant"]]
        c = row[idx["copy"]]
        if not v or not c: continue
        v = str(v).strip().upper()
        if v in ["A","B","C"]:
            out[v] = str(c).strip()
    return out

def openai_img(prompt: str, out_path: Path, api_key: str, model: str, size: str):
    url = "https://api.openai.com/v1/images/generations"
    r = requests.post(url, headers={"Authorization": f"Bearer {api_key}"},
                      json={"model": model, "prompt": prompt, "size": size},
                      timeout=180)
    if r.status_code != 200:
        raise RuntimeError(f"Images API error {r.status_code}: {r.text[:900]}")
    item = r.json()["data"][0]
    if "b64_json" in item:
        import base64
        out_path.write_bytes(base64.b64decode(item["b64_json"]))
    elif "url" in item:
        out_path.write_bytes(requests.get(item["url"], timeout=180).content)
    else:
        raise RuntimeError("Unexpected response format")

def build_prompt(season: str, kind: str, variant: str="A", mood: str="", color: str="", price: str="", offer_code: str="") -> str:
    addon = SEASON_ADDONS.get(season,"")
    oc = (offer_code or "").upper()
    offer_hint = ""
    if oc == 'SEASONPACK':
        offer_hint = 'Season pack edition cover feel, extra premium, curated seasonal atmosphere.'
    elif oc == 'D21':
        offer_hint = '21-day routine bundle feel, cohesive series tone.'
    elif oc == 'D14':
        offer_hint = '14-day bundle feel, slightly premium.'
    elif oc == 'D7':
        offer_hint = '7-day mini set feel, light and friendly.'
    # --- personal modifiers (mood/color/price) ---
    mood = (mood or "").strip()
    color = (color or "").strip()
    price_s = (price or "").strip()

    mood_map = {
        "힐링": "healing, calm, soft, cozy",
        "공감": "empathetic, warm, comforting",
        "긴급": "urgent, punchy, high-contrast but still cute",
        "프리미엄": "premium, elegant, minimal luxurious",
        "VIP": "VIP exclusive, premium, elegant",
        "축하": "celebratory, joyful, sparkling",
        "다음주": "reset, fresh start, hopeful",
    }
    mood_extra = mood_map.get(mood, mood.lower()) if mood else ""

    # Color palette hint (keeps image cohesive)
    if color:
        color_hint = f"Color palette accent: {color} pastel, harmonized with ivory background."
    else:
        color_hint = ""

    # Price tier hint → more premium composition when higher price
    premium_hint = ""
    light_hint = ""
    if price_s:
        digits = "".join(ch for ch in price_s if ch.isdigit())
        try:
            if digits and int(digits) >= 4900:
                premium_hint = "Extra premium cover feel, refined lighting, subtle luxury texture."
            else:
                light_hint = "Light, playful, friendly, simple cute composition."
        except Exception:
            pass
    if kind == "thumbnail":
        extra = {
            "A":"Premium cover composition. Balanced framing. Extra empty space for headline.",
            "B":"Premium cover composition. Slightly closer framing. Strong eye contact. Empty space for headline.",
            "C":"Premium cover composition. Wider framing. Subtle seasonal arc accent. Extra empty space for headline."
        }[variant]
        return f"{BASE_PROMPT} {addon} {extra} {offer_hint} {mood_extra} {color_hint} {premium_hint} {light_hint}"
    if kind == "cta_last_seasonpack":
        return f"{BASE_PROMPT} {addon} Ultra conversion-focused season pack landing-card layout, big clean CTA zone, bold premium headline space, limited edition feel. {offer_hint} {mood_extra} {color_hint} {premium_hint} {light_hint}"
    if kind == "cta_last":
        return f"{BASE_PROMPT} {addon} Conversion-focused clean layout, extra empty space. {offer_hint} {mood_extra} {color_hint} {premium_hint} {light_hint}"
    return f"{BASE_PROMPT} {addon} Card layout, extra empty space. {offer_hint} {mood_extra} {color_hint} {premium_hint} {light_hint}"

def wrap_lines(draw, text, font, max_w):
    text = (text or "").strip()
    if not text: return []
    if " " in text:
        words = text.split()
        lines, cur = [], ""
        for w in words:
            test = (cur + " " + w).strip()
            if draw.textlength(test, font=font) <= max_w:
                cur = test
            else:
                if cur: lines.append(cur)
                cur = w
        if cur: lines.append(cur)
        return lines
    lines, cur = [], ""
    for ch in text:
        test = cur + ch
        if draw.textlength(test, font=font) <= max_w:
            cur = test
        else:
            if cur: lines.append(cur)
            cur = ch
    if cur: lines.append(cur)
    return lines

def render_square_card(base_path: Path, out_path: Path, title: str, body: str, mood: str, color: str, font_path: Optional[str]):
    im = Image.open(base_path).convert("RGBA").resize(OUT_SQUARE, Image.LANCZOS)
    draw = ImageDraw.Draw(im)
    c = (50,44,40,255)
    f_title = pick_font(font_path, 48)
    f_meta  = pick_font(font_path, 28)
    f_body  = pick_font(font_path, 32)

    draw.text((80,110), title, font=f_title, fill=c)
    meta = " · ".join([x for x in [mood, color] if x])
    if meta:
        draw.text((80,170), meta, font=f_meta, fill=c)

    x, y, max_w = 80, 720, 920
    for para in (body or "").split("\n"):
        para = para.strip()
        if not para:
            y += f_body.size + 10
            continue
        for ln in wrap_lines(draw, para, f_body, max_w):
            draw.text((x,y), ln, font=f_body, fill=c)
            y += f_body.size + 10

    out_path.parent.mkdir(parents=True, exist_ok=True)
    im.convert("RGB").save(out_path, "PNG")

def _load_json(path: str) -> dict:
    p = Path(path)
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def _save_json(path: str, obj: dict):
    Path(path).write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

def issue_coupon_local(state_file: str, tier: str, length: int = 8) -> str:
    """
    Local random coupon issuance (offline-safe). Stores issued coupons in state_file.
    """
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    code = "".join(secrets.choice(alphabet) for _ in range(length))
    st = _load_json(state_file)
    st.setdefault("issued", {})
    st["issued"][code] = {"tier": tier, "issued_at": datetime.utcnow().isoformat(), "used": False}
    _save_json(state_file, st)
    return code

def mark_coupon_used_local(state_file: str, code: str, meta: dict | None = None):
    st = _load_json(state_file)
    if "issued" in st and code in st["issued"]:
        st["issued"][code]["used"] = True
        st["issued"][code]["used_at"] = datetime.utcnow().isoformat()
        if meta:
            st["issued"][code]["meta"] = meta
        _save_json(state_file, st)

def make_bonus_card(out_dir: Path, bonus_key: str, theme: str, font_path: str | None, preset_story: str, qr_url: str = "") -> dict:
    """
    Create square + 9:16 bonus card. Returns paths.
    bonus_key: "BONUS DAY10" / "BONUS DAY11"
    """
    title = f"{bonus_key} 전용"
    body = "구매자님 감사합니다!\n보너스 카드가 열렸어요."
    if bonus_key.endswith("11"):
        body = "VIP 혜택 오픈!\n오늘의 마음을 더 예쁘게 채워드릴게요."
    # background
    bg = Image.new("RGB", OUT_SQUARE, (252, 249, 244))
    sq = out_dir/f"{bonus_key.replace(' ','_')}.png"
    render_square_card(bg, sq, title, body, "", "", font_path)
    if theme == "gold":
        add_ribbon_badge(sq, sq, "VIP BONUS", theme="gold")
    if qr_url:
        # reuse teaser QR helper (label adjusted)
        add_teaser_qr(sq, sq, qr_url, label="다운로드")
    st = square_to_story(Image.open(sq).convert("RGB"), preset_story)
    st_path = out_dir/f"{bonus_key.replace(' ','_')}_9x16.png"
    st.save(st_path, "PNG")
    return {"square": str(sq), "story": str(st_path)}

def write_message_payload(out_dir: Path, filename: str, platform: str, tier: str, coupon_code: str, bonus_link: str, bonus_story_link: str, segment: str):
    """
# optional: send via API (Kakao AlimTalk / SMS / IG DM)
try:
    if args.send_messages and args.sender != "off":
        
# IG/TikTok: use comment+landing funnel instead of DM (more stable)
if args.funnel_mode == "comment_landing" and args.platform.lower() in ["instagram","tiktok"]:
    out_dir = Path(args.out_dir)
    payload_path = out_dir/args.message_out
    pl = json.loads(payload_path.read_text(encoding="utf-8"))
    # comment reply + pinned comment templates
    comment = build_comment_reply_payload(pl, args.platform)
    write_json(out_dir/"comment_reply_payload.json", comment)
    # landing redirect (optional)
    if args.landing_destination_url:
        landing = build_landing_payload(pl, args.landing_destination_url)
        write_json(out_dir/"landing_payload.json", landing)
        
# generate landing variants with coupon copy + optional tracking
coupon_code = pl.get("coupon_code","")
landing_files = write_landing_html_variants(
    out_dir,
    args.landing_destination_url,
    coupon_code=coupon_code,
    variants=args.landing_variants,
    track_url=args.landing_track_url,
)
profile_link_url = ""
profile_link_map = {}
if args.upload_landing:
    try:
        profile_link_map = upload_landing_variants_s3(
            landing_files,
            bucket=args.s3_bucket,
            prefix=args.landing_s3_prefix,
            public_url_base=args.s3_public_url_base,
        )
        # choose A as default profile link
        profile_link_url = profile_link_map.get("landing_A.html","")
    except Exception:
        profile_link_url = ""
        profile_link_map = {}
    send_res = {"channel":"comment_landing", "success": True, "result": {"comment_reply_payload":"comment_reply_payload.json", "landing_variants": [p.name for p in landing_files]}}
                                # write A/B report placeholder (actual comparison computed by server)
                                try:
                                    (out_dir/"landing_ab_report.json").write_text(json.dumps({"variants": list(profile_link_map.keys()), "note": "Run server_loyalty.py /report to compute conversion by variant."}, ensure_ascii=False, indent=2), encoding="utf-8")
                                except Exception:
                                    pass
else:
    send_res = dispatch_send(
        sender=args.sender,
        payload_path=Path(args.out_dir)/args.message_out,
        config_path=Path(args.sender_config),
        dry_run=args.dry_run,
        fallback_sms_on_fail=args.fallback_sms_on_fail,
    )
# log to xlsx
try:
    append_send_log_xlsx(
        Path(args.log_xlsx),
        {
            "ts": now_kst_iso(),
            "platform": args.platform,
            "segment": args.segment,
            "offer": args.offer_code,
            "tier": (bonus_info.get("tier","") if "bonus_info" in locals() else ""),
            "sender": args.sender,
            "funnel_mode": args.funnel_mode,
            "success": bool(send_res.get("success", False)),
            "channel_used": send_res.get("channel",""),
            "bonus_link": json.loads((Path(args.out_dir)/args.message_out).read_text(encoding="utf-8")).get("bonus_link",""),
            "coupon_code": json.loads((Path(args.out_dir)/args.message_out).read_text(encoding="utf-8")).get("coupon_code",""),
            "raw": json.dumps(send_res, ensure_ascii=False)[:30000],
        },
        sheet_name=args.log_sheet
    )
except Exception:
    pass
except Exception:
    pass

    Writes a JSON with message templates (for DM/알림톡/문자/메일 등 외부 발송 시스템에 그대로 전달).
    """
    # platform-specific wording
    if platform.lower() == "tiktok":
        hook = "지금 안 사면 놓쳐요 🔥"
    else:
        hook = "오늘만 열리는 혜택 🎁"
    obj = {
        "platform": platform,
        "segment": segment,
        "tier": tier,
        "hook": hook,
        "coupon_code": coupon_code,
        "bonus_link": bonus_link,
        "bonus_story_link": bonus_story_link,
        "profile_link_url": profile_link_url,
        "profile_link_map": profile_link_map,
        "message_ko": f"{hook}\n보너스 카드: {bonus_link}\n스토리용: {bonus_story_link}\n쿠폰코드: {coupon_code}",
    }
    (out_dir/filename).write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

def compute_time_left(deadline_date: str, deadline_time: str) -> tuple[int, float, float, bool]:
    """
    Returns (raw_days, hours_left, minutes_left, expired)
    """
    if not deadline_date:
        return 999, 999.0, 999.0, False
    try:
        dt = datetime.strptime(deadline_date + " " + (deadline_time or "23:59"), "%Y-%m-%d %H:%M")
        now = datetime.now()
        delta = (dt - now).total_seconds()
        minutes = delta / 60.0
        hours = minutes / 60.0
        raw_days = (dt.date() - date.today()).days
        return raw_days, max(0.0, hours), max(0.0, minutes), (delta < 0)
    except Exception:
        return 999, 999.0, 999.0, False

def urgency_stage(minutes_left: float, hours_left: float, use_urgency: bool, use_shock10: bool) -> str:
    if use_shock10 and minutes_left > 0 and minutes_left <= 1:
        return "M1"
    if use_shock10 and minutes_left > 0 and minutes_left <= 10:
        return "M10"
    if not use_urgency:
        return "NORMAL"
    if hours_left <= 1: return "H1"
    if hours_left <= 3: return "H3"
    if hours_left <= 6: return "H6"
    return "NORMAL"

def read_webhook_state_ext(state_file: str):
    try:
        p = Path(state_file)
        if not p.exists():
            return {}
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}

def read_webhook_counter(state_file: str) -> int:
    try:
        p = Path(state_file)
        if not p.exists():
            return -1
        obj = json.loads(p.read_text(encoding="utf-8"))
        # expected keys: current_buying_now, last_30min_orders
        v = obj.get("current_buying_now", obj.get("last_30min_orders", -1))
        return int(v)
    except Exception:
        return -1

def buying_now_counter() -> int:
    import time
    # pseudo-live counter (no server required)
    return 12 + int(time.time()) % 37

def decide_bonus_coupon(state: dict) -> dict:
    """
    Returns dict with keys:
      tier: light/premium/vip
      bonus: e.g., "BONUS DAY10" or "BONUS DAY11"
      coupon: e.g., "ALLOKI10"
      benefit_line: short line to append to CTA body
      theme: normal/gold
    Tiering uses last_amount and recent sums; customize freely.
    """
    try:
        last_amt = int(state.get("last_amount", 0))
        s30 = int(state.get("sum_30min", 0))
        high_recent = bool(state.get("high_amount_recent", False))
    except Exception:
        last_amt, s30, high_recent = 0, 0, False

    # VIP: very high single purchase or strong 30min revenue OR high recent flag
    if high_recent or last_amt >= 150000 or s30 >= 500000:
        return {
            "tier": "vip",
            "bonus": "BONUS DAY11",
            "coupon": "VIP10",
            "benefit_line": "💛 VIP 타임 · BONUS DAY11 + 쿠폰 VIP10",
            "theme": "gold",
        }
    # Premium: high but not VIP
    if last_amt >= 80000 or s30 >= 250000:
        return {
            "tier": "premium",
            "bonus": "BONUS DAY10",
            "coupon": "PREMIUM7",
            "benefit_line": "✨ 프리미엄 러시 · BONUS DAY10 + 쿠폰 PREMIUM7",
            "theme": "gold" if high_recent else "normal",
        }
    # Light/default
    return {
        "tier": "light",
        "bonus": "BONUS DAY09",
        "coupon": "SAVE5",
        "benefit_line": "🎁 지금 구매 · BONUS DAY09 + 쿠폰 SAVE5",
        "theme": "normal",
    }

def decide_dynamic_offer(state: dict, base_price_text: str) -> tuple[str,str]:
    """
    Returns (price_text, benefit_line).
    Rule-based (safe default). Customize thresholds as needed.
    """
    try:
        c5 = int(state.get("count_5min", 0))
        c30 = int(state.get("count_30min", 0))
        s5 = int(state.get("sum_5min", 0))
        s30 = int(state.get("sum_30min", 0))
        last_amt = int(state.get("last_amount", 0))
    except Exception:
        return base_price_text, ""

    # Tier logic:
    # - If last order is high or last 30min sum high => premium framing
    if last_amt >= 100000 or s30 >= 300000:
        return "4,900원 · 프리미엄", "VIP 구매 발생 · 보너스 혜택 강화"
    # - If momentum high => "혜택 +1" (copy only)
    if c5 >= 3 or s5 >= 80000:
        return base_price_text, "지금 구매 러시 · 보너스 혜택 진행중"
    # - Default
    return base_price_text, ""

def price_scale_from_counter(n: int) -> float:
    # 1.00 ~ 1.70
    return min(1.70, 1.00 + max(0, n-10) / 80.0)

def add_ribbon_badge(im_path: Path, out_path: Path, text: str, theme: str="normal"):
    """
    Add a commerce-style ribbon/badge at top-right.
    theme: normal=black, gold=gold.
    """
    im = Image.open(im_path).convert("RGBA")
    draw = ImageDraw.Draw(im)
    w,h = im.size
    bg = (0,0,0,180) if theme=="normal" else (212,175,55,220)  # gold
    fg = (255,255,255,255) if theme=="normal" else (30,20,0,255)
    pad = 18
    bw = int(w*0.42)
    bh = int(h*0.09)
    x1 = w - bw - pad
    y1 = pad
    draw.rounded_rectangle([x1,y1,x1+bw,y1+bh], radius=24, fill=bg)
    f = pick_font(DEFAULT_FONT, int(bh*0.42))
    draw.text((x1+24, y1+int(bh*0.22)), text, font=f, fill=fg)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    im.save(out_path, "PNG")

def add_teaser_qr(im_path: Path, out_path: Path, url: str, label: str = "알림 신청"):
    im = Image.open(im_path).convert("RGBA")
    draw = ImageDraw.Draw(im)
    if url:
        qr = qrcode.QRCode(box_size=6, border=2)
        qr.add_data(url)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB").resize((260,260), Image.NEAREST)
        xy = (OUT_SQUARE[0]-260-60, OUT_SQUARE[1]-260-80)
        im.paste(qr_img, xy)
        f = pick_font(DEFAULT_FONT, 30)
        draw.rounded_rectangle([xy[0]-20, xy[1]-60, xy[0]+260+20, xy[1]-12], radius=18, fill=(0,0,0,160))
        draw.text((xy[0]+70, xy[1]-55), label, font=f, fill=(255,255,255,255))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    im.convert("RGB").save(out_path, "PNG")

def add_qr_price_cta_square(square_path: Path, out_path: Path, qr_url: str, price_text: str, cta_text: str, font_path: Optional[str], price_scale: float=1.0):
    im = Image.open(square_path).convert("RGBA")
    draw = ImageDraw.Draw(im)

    qr = qrcode.QRCode(box_size=6, border=2)
    qr.add_data(qr_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB").resize((240,240), Image.NEAREST)
    qr_xy = (720, 700)
    im.paste(qr_img, qr_xy)

    f = pick_font(font_path, max(18, int(28*price_scale)))
    c = (50,44,40,255)
    tx, ty = qr_xy[0]-380, qr_xy[1]+30
    if price_text: draw.text((tx,ty), price_text, font=f, fill=c)
    if cta_text:   draw.text((tx,ty+44), cta_text, font=f, fill=c)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    im.convert("RGB").save(out_path, "PNG")

def square_to_story(square_rgb: Image.Image, preset: str) -> Image.Image:
    fg = square_rgb.resize((1080,1080), Image.LANCZOS)
    bg = fg.resize(OUT_STORY, Image.LANCZOS).filter(ImageFilter.GaussianBlur(18))
    canvas = bg.copy()
    if preset == "top":
        y = 90
    elif preset == "bottom":
        y = OUT_STORY[1] - 1080 - 90
    else:
        y = (OUT_STORY[1] - 1080) // 2
    canvas.paste(fg, (0,y))
    return canvas

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--season", choices=list(SEASON_ADDONS.keys()), required=True)
    ap.add_argument("--platform", choices=["instagram","tiktok"], default="instagram")
    ap.add_argument("--format", choices=["reels","shorts","post"], default="reels")
    ap.add_argument("--mode", choices=["paid","free"], default="paid")
    ap.add_argument("--offer_code", choices=["D7","D14","D21","SEASONPACK"], default="D21")
    ap.add_argument("--days", type=int, default=21, help="ignored when offer_code provided")
    ap.add_argument("--bonus", type=int, default=0, help="extra bonus cards; for SEASONPACK default=3")
    ap.add_argument("--segment", choices=["new","repeat"], default="new")
    ap.add_argument("--countdown_days", type=int, default=3, help="Fallback D-N if no deadline provided")
    ap.add_argument("--deadline", type=str, default="", help="YYYY-MM-DD; auto compute D-N for season pack CTA")
    ap.add_argument("--deadline_time", type=str, default="23:59", help="HH:MM local time; used for hour/min countdown")
    ap.add_argument("--urgency_video", action="store_true", help="swap CTA_T1 video at 6h/3h/1h before deadline")
    ap.add_argument("--shock_10min", action="store_true", help="if <=10min left, make red border + shake CTA_T1 video")
    ap.add_argument("--live_counter", action="store_true", help="at <=30min left show buying-now counter & enlarge price")
    ap.add_argument("--live_counter_source", choices=["pseudo","webhook"], default="pseudo", help="pseudo=time-based, webhook=real order count from server_webhook.py")
    ap.add_argument("--webhook_state_file", type=str, default=os.environ.get("WEBHOOK_STATE_FILE","./live_counter_state.json"), help="state json file written by webhook server")
    ap.add_argument("--high_amount_threshold", type=int, default=int(os.environ.get("HIGH_AMOUNT_THRESHOLD","50000")), help="KRW; escalate banner if order amount >= threshold")
    ap.add_argument("--dynamic_offer", action="store_true", help="adjust price/benefit in real-time based on webhook sales amounts")
    ap.add_argument("--generate_bonus_cards", action="store_true", help="generate tier-based bonus card images (e.g., DAY10/DAY11) and message payload")
    ap.add_argument("--coupon_mode", choices=["off","local_random"], default="local_random", help="coupon issuance mode")
    ap.add_argument("--coupon_state_file", type=str, default=os.environ.get("COUPON_STATE_FILE","./coupon_state.json"), help="coupon issuance/usage state file")
    ap.add_argument("--message_out", type=str, default="message_payload.json", help="output message payload json filename")
    ap.add_argument("--upload_backend", choices=["off","s3","gdrive"], default="off", help="auto upload bonus assets and inject URL into message payload")
    ap.add_argument("--require_stable_urls", action="store_true", help="require CloudFront/public URL base (no presigned URLs)")
    ap.add_argument("--send_messages", action="store_true", help="call messaging APIs after message_payload.json is created")
    ap.add_argument("--sender", choices=["off","kakao_i_alimtalk","solapi_sms","instagram_dm"], default="off", help="which sender backend to call")
    ap.add_argument("--sender_config", type=str, default=os.environ.get("SENDER_CONFIG","./sender_config.json"), help="sender config json path")
    ap.add_argument("--dry_run", action="store_true", help="do not actually send; just validate and print request payload")
    ap.add_argument("--fallback_sms_on_fail", action="store_true", help="if Kakao AlimTalk send fails, fallback to SMS (requires solapi config)")
    ap.add_argument("--log_xlsx", type=str, default=os.environ.get("LOG_XLSX","./performance_log.xlsx"), help="xlsx file to append send success/fail logs")
    ap.add_argument("--log_sheet", type=str, default=os.environ.get("LOG_SHEET","Send_Log"), help="sheet name for send logs")
    ap.add_argument("--funnel_mode", choices=["direct_send","comment_landing"], default="direct_send", help="for IG/TikTok: generate comment/landing payload instead of DM send")
    ap.add_argument("--landing_destination_url", type=str, default=os.environ.get("LANDING_DEST",""), help="final store URL for landing redirect")
    ap.add_argument("--upload_landing", action="store_true", help="upload landing.html to S3/CloudFront and inject stable profile link URL")
    ap.add_argument("--landing_s3_prefix", type=str, default=os.environ.get("LANDING_S3_PREFIX","alloki/landing/"), help="S3 prefix for landing files")
    ap.add_argument("--landing_variants", type=int, default=2, help="number of landing variants to generate (2 or 3)")
    ap.add_argument("--landing_track_url", type=str, default=os.environ.get("LANDING_TRACK_URL",""), help="optional tracking endpoint URL for landing visit events")
    ap.add_argument("--retarget_after_minutes", type=int, default=int(os.environ.get("RETARGET_AFTER_MIN","60")), help="minutes after visit to retarget if no purchase")
    ap.add_argument("--retarget_sender", choices=["off","kakao_i_alimtalk","solapi_sms"], default="off", help="sender backend for retarget messages")
    ap.add_argument("--retarget_config", type=str, default=os.environ.get("RETARGET_CONFIG","./sender_config.json"), help="sender config for retarget messages")
    ap.add_argument("--s3_bucket", type=str, default=os.environ.get("S3_BUCKET",""), help="S3 bucket name")
    ap.add_argument("--s3_prefix", type=str, default=os.environ.get("S3_PREFIX","alloki/bonus/"), help="S3 key prefix")
    ap.add_argument("--s3_public_url_base", type=str, default=os.environ.get("S3_PUBLIC_URL_BASE",""), help="If bucket is public/CloudFront, set base URL to build public URLs")
    ap.add_argument("--s3_presign_seconds", type=int, default=int(os.environ.get("S3_PRESIGN_SECONDS","604800")), help="Presigned URL expiry seconds (default 7d)")
    ap.add_argument("--gdrive_folder_id", type=str, default=os.environ.get("GDRIVE_FOLDER_ID",""), help="Google Drive folder id")
    ap.add_argument("--gdrive_service_account_json", type=str, default=os.environ.get("GDRIVE_SA_JSON",""), help="path to service account json")
    ap.add_argument("--graph_in_video", action="store_true", help="render real-time purchase graph inside CTA video when webhook data available")
    ap.add_argument("--auto_teaser", action="store_true", help="after deadline, generate next season teaser assets")
    ap.add_argument("--teaser_url", type=str, default=os.environ.get("TEASER_URL",""), help="URL for next-season notification signup QR")
    ap.add_argument("--next_season", type=str, default=os.environ.get("NEXT_SEASON",""), help="If expired, switch SEASONPACK to this season (spring/summer/autumn/winter)")
    ap.add_argument("--next_deadline", type=str, default=os.environ.get("NEXT_DEADLINE",""), help="YYYY-MM-DD for next season; used when switching")
    ap.add_argument("--cta_t1_video", action="store_true", help="generate 2s mp4 for seasonpack CTA_T1")

    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", default="Cards")
    ap.add_argument("--thumb_sheet", default="ThumbCopy")

    ap.add_argument("--thumb_pick", choices=["A","B","C","ALL"], default="ALL",
                   help="A/B/C 중 1종만 생성 또는 ALL(기본)")

    ap.add_argument("--export_story", action="store_true")
    ap.add_argument("--story_preset", choices=["top","middle","bottom"], default="middle")
    ap.add_argument("--story_last_preset", choices=["top","middle","bottom"], default="bottom")
    ap.add_argument("--story_last_preset_seasonpack", choices=["top","middle","bottom"], default="middle")

    ap.add_argument("--base_url", default="https://example.com/buy")
    ap.add_argument("--utm_campaign", default="winter_teaser")
    ap.add_argument("--font_path", default="")
    ap.add_argument("--out_dir", default="outputs")
    args = ap.parse_args()

raw_days_t, h_left, m_left, expired_time = compute_time_left(args.deadline, args.deadline_time)
# keep day-based expired from existing logic too; expired_time is more precise

    api_key = os.environ.get("OPENAI_API_KEY","").strip()
    if not api_key:
        print("ERROR: Please set OPENAI_API_KEY", file=sys.stderr)
        sys.exit(1)

# compute countdown (D-N)
cd, raw_delta, expired = compute_deadline_info(args.deadline, args.countdown_days)
cards = load_cards_xlsx(Path(args.xlsx), args.sheet)

    # offer plan
    days, bonus_n, offer_label = offer_plan(args.offer_code, args.season, args.days, args.bonus)

# auto hide / switch seasonpack if expired

if args.offer_code.upper() == "SEASONPACK" and expired:
    # After deadline: optionally generate next-season teaser immediately (no CTA)
    if args.auto_teaser and args.next_season:
        season_kr = {"spring":"봄","summer":"여름","autumn":"가을","winter":"겨울"}.get(args.next_season, args.next_season)
        teaser_sq = Path(args.out_dir)/"TEASER_SQ.png"
        teaser_st = Path(args.out_dir)/"TEASER_9x16.png"
        # simple clean teaser background
        base = Image.new("RGB", OUT_SQUARE, (250,247,242))
        render_square_card(base, teaser_sq, f"{season_kr} 시즌팩 예고", "곧 공개됩니다\n알림 받고 가장 먼저 받기", "", "", args.font)
        if args.teaser_url:
            add_teaser_qr(teaser_sq, teaser_sq, args.teaser_url, label="알림 신청")
        # story teaser (center)
        st = square_to_story(Image.open(teaser_sq).convert("RGB"), args.story_last_preset_seasonpack)
        st.save(teaser_st, "PNG")
        return

    if args.next_season:
        args.season = args.next_season
        if args.next_deadline:
            args.deadline = args.next_deadline
        # recompute countdown for next season
        cd, raw_delta, expired = compute_deadline_info(args.deadline, args.countdown_days)
    else:
        # hide season pack: downgrade to D21 (no bonus)
        args.offer_code = "D21"
        days, bonus_n, offer_label = offer_plan(args.offer_code, args.season, args.days, args.bonus)

    thumb_copy = thumb_copy_for_offer(args.offer_code, args.season)
    # allow spreadsheet override (optional)
    thumb_copy.update(load_thumb_copy_xlsx(Path(args.xlsx), args.thumb_sheet))

    out_root = Path(args.out_dir)/f"{date.today().isoformat()}_{args.season}_{args.platform}_{args.mode}_{args.offer_code}"
    ensure_dir(out_root)
    base_dir = out_root/"base_art"
    ensure_dir(base_dir)

    font_path = args.font_path if args.font_path else None

    # Thumbnails (pick)
    variants = ["A","B","C"] if args.thumb_pick=="ALL" else [args.thumb_pick]
    for v in variants:
        base = base_dir/f"THUMB_{v}_BASE.png"
        openai_img(build_prompt(args.season,"thumbnail",v, offer_code=args.offer_code), base, api_key, MODEL, API_SIZE)
        sq = Image.open(base).convert("RGB").resize(OUT_SQUARE, Image.LANCZOS)
        draw = ImageDraw.Draw(sq)
        f = pick_font(font_path, 54)
        draw.text((80,120), thumb_copy.get(v,""), font=f, fill=(50,44,40))
        sq_path = out_root/f"THUMBNAIL_{v}.png"
        sq.save(sq_path, "PNG")
        if args.export_story:
            # write bonus/coupon decision snapshot
            try:
                if args.dynamic_offer and args.live_counter_source=='webhook':
                    info_txt = Path(args.out_dir)/"BONUS_RULES.txt"
                    info_txt.write_text(
                        f"tier={bonus_info.get('tier')}\nbonus={bonus_info.get('bonus')}\ncoupon={bonus_info.get('coupon')}\ntheme={bonus_info.get('theme')}\n",
                        encoding='utf-8'
                    )
            except Exception:
                pass
# generate actual bonus card assets + message payload (tier-based)
try:
    if args.generate_bonus_cards:
        tier = bonus_info.get("tier","light")
        theme = bonus_info.get("theme","normal")
        coupon_code = ""
        if args.coupon_mode == "local_random":
            coupon_code = issue_coupon_local(args.coupon_state_file, tier)
        # decide which bonus to generate
        bonus_key = bonus_info.get("bonus","")
        bonus_link = ""
        if bonus_key in ["BONUS DAY10","BONUS DAY11"]:
            bonus_assets = make_bonus_card(Path(args.out_dir), bonus_key, theme, args.font, args.story_last_preset_seasonpack, qr_url="")
            # link placeholder (replace with S3/Drive URL in your sender)
            
# upload bonus assets and inject real URL
local_files = [Path(bonus_assets["square"]), Path(bonus_assets["story"])]
url_map = upload_bonus_assets(
    args.upload_backend,
    local_files,
    bucket=args.s3_bucket,
    prefix=args.s3_prefix,
    public_url_base=args.s3_public_url_base,
    presign_seconds=args.s3_presign_seconds,
    folder_id=args.gdrive_folder_id,
    sa_json_path=args.gdrive_service_account_json,
)
bonus_link = url_map.get(Path(bonus_assets["square"]).name, bonus_assets["square"])
        write_message_payload(Path(args.out_dir), args.message_out, args.platform, tier, coupon_code, bonus_link, url_map.get(Path(bonus_assets['story']).name, bonus_assets['story']) if 'bonus_assets' in locals() else '', args.segment)
except Exception:
    pass

          square_to_story(sq, args.story_preset).save(out_root/f"THUMBNAIL_{v}_9x16.png","PNG")

    # Day cards
    from urllib.parse import urlencode
    cards_dir = out_root/args.season.upper()
    ensure_dir(cards_dir)
    last_day = f"DAY{days:02d}"

    for i in range(1, days+1):
        day = f"DAY{i:02d}"
        info = cards.get(day, {"text":"","color":"","mood":"","price":"","cta":""})
        suffix = f"_{info.get('color','')}" if info.get("color","") else ""
        base = base_dir/f"{day}_BASE.png"
        openai_img(build_prompt(args.season,"card", "A", info.get("mood",""), info.get("color",""), info.get("price",""), offer_code=args.offer_code), base, api_key, MODEL, API_SIZE)

        square_path = cards_dir/f"{day}{suffix}.png"
        render_square_card(base, square_path, day, info.get("text",""), info.get("mood",""), info.get("color",""), font_path)

        utm = urlencode({
            "utm_source": args.platform,
            "utm_medium": "social",
            "utm_campaign": args.utm_campaign,
            "utm_content": f"{args.mode}_{day.lower()}",
        })
        qr_url = f"{args.base_url}?{utm}"

        square_to_export = square_path
        if args.mode=="free" and i!=1:
            locked = cards_dir/f"{day}{suffix}_LOCK.png"
            price = info.get("price","") or "3,900원 · 오늘만"
            cta = info.get("cta","") or ("즉시 다운로드" if args.platform=="instagram" else "지금 안 사면 놓쳐요")
            add_qr_price_cta_square(square_path, locked, qr_url, price, cta, font_path)
            square_to_export = locked

        if args.export_story:
            # write bonus/coupon decision snapshot
            try:
                if args.dynamic_offer and args.live_counter_source=='webhook':
                    info_txt = Path(args.out_dir)/"BONUS_RULES.txt"
                    info_txt.write_text(
                        f"tier={bonus_info.get('tier')}\nbonus={bonus_info.get('bonus')}\ncoupon={bonus_info.get('coupon')}\ntheme={bonus_info.get('theme')}\n",
                        encoding='utf-8'
                    )
            except Exception:
                pass

            sq_img = Image.open(square_to_export).convert("RGB")
            preset = args.story_last_preset if day==last_day else args.story_preset
            story = square_to_story(sq_img, preset)
            story.save(cards_dir/square_to_export.name.replace(".png","_9x16.png"), "PNG")

            if day==last_day:
                # Dedicated CTA cut
                cta_base = base_dir/f"{day}_CTA_BASE.png"
                cta_kind = "cta_last_seasonpack" if args.offer_code.upper()=="SEASONPACK" else "cta_last"
                openai_img(build_prompt(args.season, cta_kind, "A", info.get("mood",""), info.get("color",""), info.get("price",""), offer_code=args.offer_code), cta_base, api_key, MODEL, API_SIZE)
                cta_sq = cards_dir/f"{day}{suffix}_CTA.png"
                
# CTA (story last cut) – SEASONPACK 2-step CTA
if args.offer_code.upper() == "SEASONPACK":
    title, body, price, cta = seasonpack_cta_copy(args.platform, args.season, cd, args.segment)
            benefit_line = ""

bonus_info = {"tier":"light","bonus":"","coupon":"","benefit_line":"","theme":"normal"}
if args.dynamic_offer and args.live_counter and args.live_counter_source=='webhook':
    bonus_info = decide_bonus_coupon(state)
            if args.dynamic_offer and args.live_counter and args.live_counter_source=='webhook':
                price, benefit_line = decide_dynamic_offer(state, price)
                if benefit_line:
                    body = body + "\n" + benefit_line
                if bonus_info.get('benefit_line'):
                    body = body + "\n" + bonus_info['benefit_line']

    # CTA Step 1: Teaser
    cta1_base = cards_dir/f"{day}{suffix}_CTA_T1_BASE.png"
    openai_img(build_prompt(args.season,"cta_last_seasonpack","A", info.get("mood",""), info.get("color",""), info.get("price",""), offer_code=args.offer_code),
              cta1_base, api_key, MODEL, API_SIZE)
    cta1_sq = cards_dir/f"{day}{suffix}_CTA_T1.png"
    render_square_card(cta1_base, cta1_sq, title, seasonpack_cta_t1_teaser_by_stage(cd, args.platform, args.segment), info.get("mood",""), info.get("color",""), font_path)
    badge1 = "오늘 마감" if cd <= 0 else ("내일 마감" if cd <= 1 else "LIMITED")
            add_commerce_badge(cta1_sq, cta1_sq, badge1, ribbon=True)
    square_to_story(Image.open(cta1_sq).convert("RGB"), args.story_last_preset_seasonpack)                .save(cards_dir/f"{day}{suffix}_CTA_T1_9x16.png","PNG")

    # CTA Step 2: Conversion
    cta2_base = cards_dir/f"{day}{suffix}_CTA_T2_BASE.png"
    openai_img(build_prompt(args.season,"cta_last_seasonpack","A", info.get("mood",""), info.get("color",""), info.get("price",""), offer_code=args.offer_code),
              cta2_base, api_key, MODEL, API_SIZE)
    cta2_sq = cards_dir/f"{day}{suffix}_CTA_T2.png"
    render_square_card(cta2_base, cta2_sq, title, body, info.get("mood",""), info.get("color",""), font_path)
    
live_n = None
scale = 1.0
if args.live_counter and m_left <= 30:
    
state = read_webhook_state_ext(args.webhook_state_file) if args.live_counter_source=='webhook' else {}
live_30 = int(state.get("count_30min", -1))
live_5 = int(state.get("count_5min", -1))
high_hit = bool(state.get("high_amount_hit", False))
            high_recent = bool(state.get("high_amount_recent", False))
            bins30 = state.get("bins_30min", [])
if live_30 < 0:
    live_30 = buying_now_counter()
if live_5 < 0:
    live_5 = max(1, live_30//3)
live_n = live_30
                if live_n < 0:
                    live_n = buying_now_counter()
    scale = price_scale_from_counter(live_n)
add_qr_price_cta_square(cta2_sq, cta2_sq, qr_url, price, cta, font_path, price_scale=scale)
            # ribbon badge (coupon/bonus)
            if bonus_info.get('coupon'):
                add_ribbon_badge(cta2_sq, cta2_sq, f"쿠폰 {bonus_info['coupon']}", theme=bonus_info.get('theme','normal'))
    badge2 = "마지막 기회" if cd <= 0 else ("곧 마감" if cd <= 3 else "BEST VALUE")
            add_commerce_badge(cta2_sq, cta2_sq, badge2, ribbon=True)
    square_to_story(Image.open(cta2_sq).convert("RGB"), args.story_last_preset_seasonpack)                .save(cards_dir/f"{day}{suffix}_CTA_T2_9x16.png","PNG")

            # countdown label on CTA_T2 square + story
            add_countdown_label(cta2_sq, cta2_sq, cd)
            add_countdown_label(cards_dir/f"{day}{suffix}_CTA_T2_9x16.png", cards_dir/f"{day}{suffix}_CTA_T2_9x16.png", cd)

# live counter overlay (<=30min)
if args.live_counter and m_left <= 30:
    n2 = live_n if live_n is not None else (read_webhook_counter(args.webhook_state_file) if args.live_counter_source=='webhook' else buying_now_counter())
                if n2 < 0:
                    n2 = buying_now_counter()
    imx = Image.open(cards_dir/f"{day}{suffix}_CTA_T2_9x16.png").convert("RGBA")
    drawx = ImageDraw.Draw(imx)
    msg = f"최근 5분 {live_5}명 · 30분 {live_30}명 구매 중"
    fnt = pick_font(DEFAULT_FONT, int(OUT_STORY[1]*0.045))
    drawx.rounded_rectangle([60, OUT_STORY[1]-170, OUT_STORY[0]-60, OUT_STORY[1]-90], radius=26, fill=(0,0,0,180))
    drawx.text((90, OUT_STORY[1]-160), msg, font=fnt, fill=(255,255,255,255))
    imx.save(cards_dir/f"{day}{suffix}_CTA_T2_9x16.png","PNG")
            if bonus_info.get('coupon'):
                add_ribbon_badge(cards_dir/f"{day}{suffix}_CTA_T2_9x16.png", cards_dir/f"{day}{suffix}_CTA_T2_9x16.png", f"쿠폰 {bonus_info['coupon']}", theme=bonus_info.get('theme','normal'))

            # CTA_T1 mp4 (optional)
            
if args.cta_t1_video:
    stage = urgency_stage(m_left, h_left, args.urgency_video, args.shock_10min)
    mp4_path = cards_dir/f"{day}{suffix}_CTA_T1_9x16_{stage}.mp4"
    make_cta_t1_mp4(
        cards_dir/f"{day}{suffix}_CTA_T1_9x16.png",
        mp4_path,
        seconds=2.0,
        fps=30,
        shake=(stage=="M10"),
        red_border=(stage=="M10"),
    )

else:
    pass(cards_dir/f"{day}{suffix}_CTA_9x16.png","PNG")

# Bonus cards (SEASONPACK)
if args.offer_code.upper() == "SEASONPACK" and bonus_n > 0:
    for j in range(1, bonus_n+1):
        bkey = f"BONUS{j:02d}"
        info = cards.get(bkey, {"text": f"보너스 카드 {j:02d} · 시즌팩 구매자 전용", "color":"", "mood":"프리미엄", "price":"", "cta":""})
        suffix = f"_{info.get('color','')}" if info.get("color","") else ""
        base = base_dir/f"{bkey}_BASE.png"
        openai_img(build_prompt(args.season,"card","A", info.get("mood",""), info.get("color",""), info.get("price",""), offer_code=args.offer_code),
                  base, api_key, MODEL, API_SIZE)

        square_path = cards_dir/f"{bkey}{suffix}.png"
        render_square_card(base, square_path, bkey, info.get("text",""), info.get("mood",""), info.get("color",""), font_path)

        utm = urlencode({
            "utm_source": args.platform,
            "utm_medium": "social",
            "utm_campaign": args.utm_campaign,
            "utm_content": f"{args.mode}_{bkey.lower()}",
        })
        qr_url = f"{args.base_url}?{utm}"

        square_to_export = square_path
        if args.mode=="free":
            locked = cards_dir/f"{bkey}{suffix}_LOCK.png"
            price = info.get("price","") or "12,900원 · 시즌팩"
            cta = info.get("cta","") or ("즉시 다운로드" if args.platform=="instagram" else "지금 안 사면 놓쳐요")
            add_qr_price_cta_square(square_path, locked, qr_url, price, cta, font_path)
            square_to_export = locked

        if args.export_story:
            # write bonus/coupon decision snapshot
            try:
                if args.dynamic_offer and args.live_counter_source=='webhook':
                    info_txt = Path(args.out_dir)/"BONUS_RULES.txt"
                    info_txt.write_text(
                        f"tier={bonus_info.get('tier')}\nbonus={bonus_info.get('bonus')}\ncoupon={bonus_info.get('coupon')}\ntheme={bonus_info.get('theme')}\n",
                        encoding='utf-8'
                    )
            except Exception:
                pass

            sq_img = Image.open(square_to_export).convert("RGB")
            story = square_to_story(sq_img, args.story_preset)
            story.save(cards_dir/square_to_export.name.replace(".png","_9x16.png"), "PNG")

    # ZIP
    zip_path = out_root.with_suffix(".zip")
    with zipfile.ZipFile(zip_path,"w",zipfile.ZIP_DEFLATED) as z:
        for p in sorted(out_root.rglob("*")):
            if p.is_file():
                z.write(p, arcname=str(p.relative_to(out_root)))
    print("DONE:", zip_path)

if __name__ == "__main__":
    main()from send_dispatch import dispatch_send
from log_to_sheet import append_send_log_xlsx, now_kst_iso
from funnel_tools import build_comment_reply_payload, build_landing_payload, write_json, write_landing_html, write_landing_html_variants
from uploaders import upload_bonus_assets, upload_landing_html_s3, upload_landing_variants_s3

