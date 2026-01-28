#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Alloki & Dalloki Unified Suite (All-in-One Entry Point)

This app.py is a stable orchestrator that wraps the existing modules/scripts in this package:
- run_generate.py : generate cards/thumbs/story + (optional) upload/send/log
- server_webhook_platforms.py : webhook receiver / live counters (optional background)
- update_ltv : update platform_ev_config.json with data-driven LTV weights

Key goal: **one program** + **one-click commands** like:
  python app.py run_week --season spring --platforms instagram,tiktok --segments new,repeat --auto_server

Any extra args after `--` are forwarded to run_generate.py so you can keep extending without breaking the entrypoint.
Example:
  python app.py run_week --season promo_d-3 --platforms tiktok --segments new -- --shock_10min --urgency_video
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import openpyxl

ROOT = Path(__file__).resolve().parent


def _run(cmd: List[str], cwd: Path | None = None) -> int:
    p = subprocess.run(cmd, cwd=str(cwd) if cwd else None)
    return int(p.returncode)


def _popen(cmd: List[str], cwd: Path | None = None):
    return subprocess.Popen(cmd, cwd=str(cwd) if cwd else None, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)


def _today_iso() -> str:
    import datetime
    return datetime.date.today().isoformat()


def _monday(d):
    import datetime
    return d - datetime.timedelta(days=d.weekday())


def _date_from_iso(s: str):
    import datetime
    return datetime.date.fromisoformat(s)


def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


# ----------------------------
# LTV update (data-driven)
# ----------------------------

def _to_float(x):
    try:
        return float(x)
    except Exception:
        return None


def _clamp(x: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, x))


def _load_ev_config(path: Path) -> dict:
    if not path.exists():
        return {"default": {}}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {"default": {}}


def _iter_rows_with_headers(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        header_row = None
        headers = None
        for r in range(1, 21):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            lower = [str(x).strip().lower() if x is not None else "" for x in row]
            if "platform" in lower and ("conv_rate" in lower or "cvr" in lower or "conversion" in lower):
                header_row = r
                headers = lower
                break
        if header_row is None:
            continue
        col = {h: i + 1 for i, h in enumerate(headers) if h}
        for r in range(header_row + 1, ws.max_row + 1):
            if all(ws.cell(r, c).value in (None, "") for c in range(1, ws.max_column + 1)):
                continue
            yield ws.title, col, ws, r


def _get_cell(ws, r, colmap, key, default=None):
    aliases = {
        "platform": ["platform", "channel", "source_platform"],
        "season": ["season", "campaign", "promo", "season_label"],
        "segment": ["segment", "audience", "cohort"],
        "conv_rate": ["conv_rate", "cvr", "conversion_rate", "links_cvr", "metric"],

        "repurchase_rate": ["repurchase_rate", "repeat_rate", "repurchase_cvr"],
        "repurchase": ["repurchase", "repeat", "repeat_conversions"],
        "unique_clickers": ["unique_clickers", "unique_clicker", "uniq_clickers", "uniq_clicker"],

        "coupon_use_rate": ["coupon_use_rate", "coupon_rate", "coupon_redeem_rate"],
        "coupon_used": ["coupon_used", "coupon_redeemed", "coupon_count"],
        "conversions": ["conversions", "orders", "purchases", "buy"],

        "month": ["month", "yyyymm", "period"],

        "revenue": ["revenue", "sales", "gmv", "order_amount"],
        "repurchase_revenue": ["repurchase_revenue", "repeat_revenue", "ltv_revenue"],
        "coupon_discount_rate": ["coupon_discount_rate", "discount_rate", "coupon_rate_pct", "coupon_pct"],

        "product_type": ["product_type", "product", "offer_type"],
    }
    keys = aliases.get(key, [key])
    for k in keys:
        if k in colmap:
            v = ws.cell(r, colmap[k]).value
            if v is None or v == "":
                continue
            return v
    return default


def update_ltv_from_xlsx(xlsx_path: Path, cfg_path: Path, month: str = "") -> dict:
    cfg = _load_ev_config(cfg_path)
    cfg.setdefault("default", {})

    model = cfg["default"].get("ltv_model", {})
    alpha = float(model.get("alpha_repurchase_rev_ratio", 0.8))
    beta  = float(model.get("beta_coupon_discount", 0.5))
    gamma = float(model.get("gamma_seasonpack_uplift", 0.15))

    agg: Dict[str, Dict[str, List[float]]] = {}
    seg_perf: Dict[str, List[float]] = {"new": [], "repeat": []}

    for _, colmap, ws, r in _iter_rows_with_headers(xlsx_path):
        plat = _get_cell(ws, r, colmap, "platform")
        if plat is None:
            continue
        plat = str(plat).strip().lower()

        if month:
            mcol = _get_cell(ws, r, colmap, "month")
            if mcol is not None and str(mcol).strip() != str(month).strip():
                continue

        if plat not in agg:
            agg[plat] = {
                "repurchase_rate": [],
                "coupon_use_rate": [],
                "revenue": [],
                "repurchase_revenue": [],
                "coupon_discount_rate": [],
                "seasonpack_flag": [],
            }

        rr = _to_float(_get_cell(ws, r, colmap, "repurchase_rate"))
        if rr is None:
            rep = _to_float(_get_cell(ws, r, colmap, "repurchase"))
            uc = _to_float(_get_cell(ws, r, colmap, "unique_clickers"))
            if rep is not None and uc and uc > 0:
                rr = rep / uc
        if rr is not None:
            agg[plat]["repurchase_rate"].append(float(rr))

        cur = _to_float(_get_cell(ws, r, colmap, "coupon_use_rate"))
        if cur is None:
            cu = _to_float(_get_cell(ws, r, colmap, "coupon_used"))
            conv = _to_float(_get_cell(ws, r, colmap, "conversions"))
            if cu is not None and conv and conv > 0:
                cur = cu / conv
        if cur is not None:
            agg[plat]["coupon_use_rate"].append(float(cur))

        rev = _to_float(_get_cell(ws, r, colmap, "revenue"))
        rep_rev = _to_float(_get_cell(ws, r, colmap, "repurchase_revenue"))
        disc = _to_float(_get_cell(ws, r, colmap, "coupon_discount_rate"))
        if disc is not None and disc > 1.0:
            disc = disc / 100.0

        if rev is not None:
            agg[plat]["revenue"].append(float(rev))
        if rep_rev is not None:
            agg[plat]["repurchase_revenue"].append(float(rep_rev))
        if disc is not None:
            agg[plat]["coupon_discount_rate"].append(float(disc))

        ptype = _get_cell(ws, r, colmap, "product_type", default="standard")
        agg[plat]["seasonpack_flag"].append(1.0 if str(ptype).strip().lower() == "seasonpack" else 0.0)

        seg = str(_get_cell(ws, r, colmap, "segment", default="")).strip().lower()
        cvr = _to_float(_get_cell(ws, r, colmap, "conv_rate"))
        if seg in seg_perf and cvr is not None:
            seg_perf[seg].append(float(cvr))

    # platform ltv_weight
    for plat, vals in agg.items():
        rr_avg = sum(vals["repurchase_rate"]) / len(vals["repurchase_rate"]) if vals["repurchase_rate"] else None
        cur_avg = sum(vals["coupon_use_rate"]) / len(vals["coupon_use_rate"]) if vals["coupon_use_rate"] else None

        base = 1.0
        if vals["revenue"] and vals["repurchase_revenue"] and sum(vals["revenue"]) > 0:
            rep_ratio = sum(vals["repurchase_revenue"]) / max(1.0, sum(vals["revenue"]))
            base += alpha * float(rep_ratio)
        elif rr_avg is not None:
            base += 0.6 * float(rr_avg)

        if vals["coupon_discount_rate"]:
            avg_disc = sum(vals["coupon_discount_rate"]) / len(vals["coupon_discount_rate"])
            base -= beta * float(avg_disc)
        elif cur_avg is not None:
            base += 0.2 * float(cur_avg)

        sp_share = sum(vals["seasonpack_flag"]) / len(vals["seasonpack_flag"]) if vals["seasonpack_flag"] else 0.0
        base += gamma * float(sp_share)

        ltv = _clamp(base, 0.8, 2.5)
        cfg.setdefault(plat, {})
        if isinstance(cfg[plat], dict):
            cfg[plat]["ltv_weight"] = round(ltv, 4)

    # repeat multiplier
    try:
        new_avg = sum(seg_perf["new"]) / len(seg_perf["new"]) if seg_perf["new"] else None
        rep_avg = sum(seg_perf["repeat"]) / len(seg_perf["repeat"]) if seg_perf["repeat"] else None
        if new_avg and rep_avg and new_avg > 0:
            mult = _clamp(rep_avg / new_avg, 1.0, 1.6)
            cfg["default"].setdefault("segment_ltv_mult", {})
            if isinstance(cfg["default"]["segment_ltv_mult"], dict):
                cfg["default"]["segment_ltv_mult"]["repeat"] = round(mult, 4)
    except Exception:
        pass

    cfg_path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    return cfg


# ----------------------------
# One-click pipeline wrappers
# ----------------------------

def cmd_update_ltv(args) -> int:
    xlsx = Path(args.xlsx)
    cfgp = Path(args.platform_ev_config)
    update_ltv_from_xlsx(xlsx, cfgp, month=args.month)
    print("Updated:", cfgp)
    return 0


def cmd_run_week(args) -> int:
    platforms = [p.strip() for p in args.platforms.split(",") if p.strip()]
    segments = [s.strip() for s in args.segments.split(",") if s.strip()]
    out_root = Path(args.out_dir)
    _ensure_dir(out_root)

    # optional background server
    server_proc = None
    if args.auto_server:
        server_cmd = [sys.executable, str(ROOT/"server_webhook_platforms.py"), "--port", str(args.server_port)]
        server_proc = _popen(server_cmd, cwd=ROOT)
        time.sleep(0.8)  # give it a moment

    # forward extra args to run_generate.py
    forward = args.forward_args or []

    rc = 0
    for plat in platforms:
        for seg in segments:
            out_dir = out_root / f"{args.season}" / plat / seg
            _ensure_dir(out_dir)

            cmd = [
                sys.executable, str(ROOT/"run_generate.py"),
                "--season", args.season,
                "--platform", plat,
                "--segment", seg,
                "--days", str(args.days),
                "--format", args.format,
                "--mode", args.mode,
                "--offer_code", args.offer_code,
                "--message_out", str(out_dir/"message_payload.json"),
                "--log_xlsx", str(Path(args.log_xlsx) if args.log_xlsx else out_root/"performance_log.xlsx"),
                "--platform_ev_config", str(args.platform_ev_config) if args.platform_ev_config else str(ROOT/"platform_ev_config.json"),
            ]

            if args.deadline:
                cmd += ["--deadline", args.deadline]
            if args.deadline_time:
                cmd += ["--deadline_time", args.deadline_time]
            if args.upload_backend:
                cmd += ["--upload_backend", args.upload_backend]
            if args.require_stable_urls:
                cmd += ["--require_stable_urls"]
            if args.send_messages:
                cmd += ["--send_messages", "--sender", args.sender, "--sender_config", args.sender_config]
            if args.dry_run:
                cmd += ["--dry_run"]

            # pass-through args after --
            if forward:
                cmd += forward

            print("\n[RUN]", " ".join(cmd))
            rc = _run(cmd, cwd=ROOT)
            if rc != 0:
                print("❌ run_generate failed for", plat, seg, "rc=", rc)
                break
        if rc != 0:
            break

    if server_proc:
        try:
            server_proc.terminate()
            server_proc.wait(timeout=5)
        except Exception:
            try:
                server_proc.kill()
            except Exception:
                pass

    return int(rc)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Alloki & Dalloki Unified Suite (One Program)")
    sub = p.add_subparsers(dest="cmd", required=True)

    rw = sub.add_parser("run_week", help="One-click: generate this week's sets (platform x segment) via run_generate.py")
    rw.add_argument("--season", required=True, help="e.g., spring / yearend / promo_d-3")
    rw.add_argument("--platforms", default="instagram,tiktok", help="comma list")
    rw.add_argument("--segments", default="new,repeat", help="comma list")
    rw.add_argument("--days", type=int, default=7, help="card count (7/14/21). Note: 21 usually seasonpack.")
    rw.add_argument("--offer_code", default="D7", help="D7/D14/D21/SEASONPACK... forwarded to run_generate")
    rw.add_argument("--format", default="both", help="square|story|both")
    rw.add_argument("--mode", default="paid", help="paid|free-lock etc (depends on run_generate)")
    rw.add_argument("--out_dir", default="./out_week", help="output root")
    rw.add_argument("--log_xlsx", default="./performance_log.xlsx", help="performance log xlsx")
    rw.add_argument("--platform_ev_config", default="./platform_ev_config.json", help="EV/LTV config json")

    rw.add_argument("--deadline", default="", help="YYYY-MM-DD (optional)")
    rw.add_argument("--deadline_time", default="", help="HH:MM (optional)")
    rw.add_argument("--auto_server", action="store_true", help="start webhook server in background during generation")
    rw.add_argument("--server_port", type=int, default=8787)

    rw.add_argument("--upload_backend", default="", help="s3|gdrive|none")
    rw.add_argument("--require_stable_urls", action="store_true")
    rw.add_argument("--send_messages", action="store_true")
    rw.add_argument("--sender", default="kakao", help="kakao|sms")
    rw.add_argument("--sender_config", default=str(ROOT/"sender_config.json"))
    rw.add_argument("--dry_run", action="store_true")

    # Everything after -- will be forwarded to run_generate.py
    rw.add_argument("forward_args", nargs=argparse.REMAINDER, help="Use `-- <args...>` to pass through to run_generate.py")
    rw.set_defaults(func=cmd_run_week)

    ul = sub.add_parser("update_ltv", help="Update platform_ev_config.json using real data from performance xlsx")
    ul.add_argument("--xlsx", required=True)
    ul.add_argument("--platform_ev_config", default="./platform_ev_config.json")
    ul.add_argument("--month", default="", help="optional: YYYY-MM")
    ul.set_defaults(func=cmd_update_ltv)

    return p


def main(argv: List[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
