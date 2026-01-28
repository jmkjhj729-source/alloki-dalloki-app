from __future__ import annotations
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any

def append_send_log_xlsx(xlsx_path: Path, row: Dict[str, Any], sheet_name: str="Send_Log") -> None:
    """
    Append a row to an Excel workbook (create if missing).
    Columns are created from keys (stable order).
    """
    from openpyxl import Workbook, load_workbook

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        # remove default sheet
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(list(row.keys()))

    # if header missing, create
    if ws.max_row == 0:
        ws.append(list(row.keys()))

    # ensure header matches (best-effort)
    header = [c.value for c in ws[1]]
    if header != list(row.keys()):
        # align to existing header
        aligned = [row.get(k, "") for k in header]
        ws.append(aligned)
    else:
        ws.append([row.get(k, "") for k in row.keys()])

    wb.save(xlsx_path)

def now_kst_iso() -> str:
    # naive ISO; you can replace with timezone-aware if needed
    return datetime.now().isoformat(timespec="seconds")
