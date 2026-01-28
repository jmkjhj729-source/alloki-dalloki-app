#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
v16 – 요일·플랫폼 최적 mood/color 자동 + price A/B 분기 + 프롬프트 톤 분기(진짜 '분위기'가 바뀜)

✅ 구현 방식(안전/확실)
1) 서버가 구매자별로 mood/color/price를 결정
2) CARDS_XLSX를 임시 복사 → DAY09/DAY10 행의 (mood,color,price,cta) 값을 buyer용으로 overwrite
3) patched run_generate.py가 그 값을 읽어서
   - 이미지 생성 프롬프트(build_prompt)에 mood/color/price 힌트까지 포함
4) 생성된 베이스 이미지가 이미 '분위기'가 바뀐 상태에서 카드가 만들어짐
5) 그 위에 프리셋(top/middle/bottom) 텍스트/배지 오버레이까지 적용

필수 env
- OPENAI_API_KEY
- TRACKER_XLSX
- CARDS_XLSX

선택 env
- BONUS_FONT_PATH (한글 폰트 권장)
- RUN_GENERATE_PATH (기본: 같은 폴더 run_generate.py)
"""

from __future__ import annotations
import argparse, json, os, secrets, time, sqlite3, subprocess, sys, shutil, threading
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, Optional

from fastapi import FastAPI, Request
from fastapi.responses import RedirectResponse, JSONResponse
import uvicorn
import openpyxl
from PIL import Image, ImageDraw, ImageFont

APP = FastAPI()

DB_PATH = Path(os.environ.get("PROFILE_DB", "buyer_profile.sqlite"))
TRACKER_XLSX = Path(os.environ.get("TRACKER_XLSX", "Alloki_Dalloki_Performance_Tracker.xlsx"))
CARDS_XLSX = Path(os.environ.get("CARDS_XLSX", "day_texts.xlsx"))

RUN_GENERATE = Path(os.environ.get("RUN_GENERATE_PATH", "run_generate.py"))
BONUS_SEASON = os.environ.get("BONUS_SEASON", "winter")
BONUS_OUT_DIR = Path(os.environ.get("BONUS_OUT_DIR", "bonus_out"))
FONT_PATH = os.environ.get("BONUS_FONT_PATH", "")
PRESETS_PATH = Path(os.environ.get("OVERLAY_PRESETS", "overlay_presets.json"))
DEFAULT_PRESET_INSTAGRAM = os.environ.get("DEFAULT_PRESET_INSTAGRAM", "top")
DEFAULT_PRESET_TIKTOK = os.environ.get("DEFAULT_PRESET_TIKTOK", "middle")


# ---------- Decision Metric per Platform ----------
# Which EV to use when selecting price variant by platform
# instagram/tiktok -> click-based EV
# smartstore/web   -> link-based EV

# ---------- Multi-dimension EV weights (platform × weekday × season) ----------
# weights multiply into EV when selecting variants/offers.
# default is 1.0 if missing.
PLATFORM_EV_WEIGHT = {
    "instagram": 1.0, "tiktok": 1.0, "smartstore": 1.0, "web": 1.0, "store": 1.0
}
# Monday..Sunday in Korean single char: 월화수목금토일
WEEKDAY_EV_WEIGHT = {"월": 1.0, "화": 1.0, "수": 1.0, "목": 1.0, "금": 1.0, "토": 1.0, "일": 1.0}
SEASON_EV_WEIGHT = {"spring": 1.0, "summer": 1.0, "autumn": 1.0, "winter": 1.0}

def weighted_ev(platform: str, weekday: str, season: str, ev_base: float) -> float:
    p = PLATFORM_EV_WEIGHT.get((platform or "").lower(), 1.0)
    w = WEEKDAY_EV_WEIGHT.get((weekday or "").strip(), 1.0)
    s = SEASON_EV_WEIGHT.get((season or "").lower(), 1.0)
    return float(ev_base) * float(p) * float(w) * float(s)

PLATFORM_DECISION_METRIC = {
    "instagram": "ev_clickers",
    "tiktok": "ev_clickers",
    "reels": "ev_clickers",
    "shorts": "ev_clickers",
    "smartstore": "ev_links",
    "store": "ev_links",
    "web": "ev_links",
}



# ---------- Offer catalog (days / packs) ----------
# offer_code examples: "D7","D14","D21","SEASONPACK"
DEFAULT_OFFER_CODES = ["D7","D14","D21","SEASONPACK"]

OFFER_PRICE_MAP = {
    "D7": int(os.environ.get("OFFER_PRICE_7", "3900")),
    "D14": int(os.environ.get("OFFER_PRICE_14", "4900")),
    "D21": int(os.environ.get("OFFER_PRICE_21", "7900")),
    "SEASONPACK": int(os.environ.get("OFFER_PRICE_SEASONPACK", "12900")),
}

def offer_code_to_days(code: str) -> int:
    code = (code or "").upper()
    if code == "D7": return 7
    if code == "D14": return 14
    if code == "D21": return 21
    return 0


# ---------- Offer selection (7/14/21/SeasonPack) ----------
def read_offer_stats(tracker: Path, segment: str, platform: str, wday: str, season: str) -> Optional[Dict[str, Any]]:
    """
    Sheet: Offer_Stats
    Columns (recommended):
      segment | platform | weekday | season | offer_code | offer_days | price | conv_rate_links | click_cvr | ev_links | ev_clickers
    offer_code is preferred. If missing, offer_days (7/14/21) is used.
    """
    if not tracker.exists():
        return None
    try:
        wb = openpyxl.load_workbook(tracker)
        if "Offer_Stats" not in wb.sheetnames:
            return None
        ws = wb["Offer_Stats"]
        headers = [safe_str(c.value).lower() for c in ws[1]]
        idxm = {h:i for i,h in enumerate(headers) if h}

        def get(row, name, default=None):
            i = idxm.get(name, None)
            return row[i] if i is not None and i < len(row) else default

        metric = PLATFORM_DECISION_METRIC.get(platform.lower(), "ev_links")
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if safe_str(get(r,"segment","")).lower() != segment.lower():
                continue
            if safe_str(get(r,"platform","")).lower() != platform.lower():
                continue
            if safe_str(get(r,"weekday","")) != wday:
                continue
            if safe_str(get(r,"season","")).lower() != season.lower():
                continue

            offer_code = safe_str(get(r,"offer_code","")).upper()
            offer_days = int(float(get(r,"offer_days",0) or 0))
            if not offer_code:
                offer_code = "D21" if offer_days==21 else ("D14" if offer_days==14 else ("D7" if offer_days==7 else ""))
            if not offer_code:
                continue

            price = int(float(get(r,"price", OFFER_PRICE_MAP.get(offer_code,0)) or 0))
            ev_links = float(get(r,"ev_links",0.0) or 0.0)
            ev_clickers = float(get(r,"ev_clickers",0.0) or 0.0)
            base = ev_clickers if metric=="ev_clickers" else ev_links

            rows.append({
                "offer_code": offer_code,
                "offer_days": offer_code_to_days(offer_code),
                "price": price,
                "ev_links": ev_links,
                "ev_clickers": ev_clickers,
                "weighted_ev": weighted_ev(platform, wday, season, base),
            })
        if not rows:
            return None
        return sorted(rows, key=lambda x: x["weighted_ev"], reverse=True)[0]
    except Exception:
        return None

def choose_offer(buyer_id: str, segment: str, platform: str, wday: str, season: str) -> Tuple[str, int]:
    """
    Returns (offer_code, offer_days)
    1) Offer_Stats best weighted EV
    2) fallback heuristic:
       - 일요일: SEASONPACK (부드럽게 업셀)
       - repeat: 금/토/일 -> D14, 월~목 -> D21
       - new: D7
    """
    best = read_offer_stats(TRACKER_XLSX, segment, platform, wday, season)
    if best:
        return best["offer_code"], int(best["offer_days"] or 0)
    if wday == "일":
        return "SEASONPACK", 0
    if segment == "repeat":
        if wday in ("금","토","일"):
            return "D14", 14
        return "D21", 21
    return "D7", 7

AUTO_MONTHLY_STATS = os.environ.get("AUTO_MONTHLY_STATS", "1").strip() == "1"
MONTHLY_STATS_HOUR = int(os.environ.get("MONTHLY_STATS_HOUR", "3"))  # 03:00 local time
MONTHLY_STATS_MINUTE = int(os.environ.get("MONTHLY_STATS_MINUTE", "10"))
STATS_LOOKBACK_MONTHS = int(os.environ.get("STATS_LOOKBACK_MONTHS", "1"))  # 기본: 직전 1개월
CONV_WINDOW_DAYS = int(os.environ.get("CONV_WINDOW_DAYS", "7"))  # 클릭 후 구매 전환 윈도우
TIMEZONE = os.environ.get("TIMEZONE", "Asia/Seoul")


def safe_str(x): return "" if x is None else str(x).strip()

def load_presets() -> Dict[str, Any]:
    if PRESETS_PATH.exists():
        return json.loads(PRESETS_PATH.read_text(encoding="utf-8"))
    return {}
PRESETS = load_presets()

# ---------- DB ----------
def db():
    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA journal_mode=WAL;")
    return con

def init_db():
    con = db()
    con.execute("""
    CREATE TABLE IF NOT EXISTS buyers(
        buyer_id TEXT PRIMARY KEY,
        buyer_name TEXT,
        created_at REAL
    )""")
    con.execute("""
    CREATE TABLE IF NOT EXISTS events(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        buyer_id TEXT,
        event_type TEXT,
        platform TEXT,
        order_id TEXT,
        product_name TEXT,
        created_at REAL
    )""")
    con.execute("""
    CREATE TABLE IF NOT EXISTS bonus_links(
        token TEXT PRIMARY KEY,
        buyer_id TEXT,
        day TEXT,
        target_url TEXT,
        platform TEXT,
        created_at REAL,
        clicks INTEGER DEFAULT 0
    )""")
    con.execute("""
    CREATE TABLE IF NOT EXISTS clicks(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        token TEXT,
        buyer_id TEXT,
        day TEXT,
        platform TEXT,
        ts REAL,
        ua TEXT,
        ref TEXT
    )""")

    # Schema evolution (safe ALTER)
    for stmt in [
        "ALTER TABLE bonus_links ADD COLUMN season TEXT",
        "ALTER TABLE bonus_links ADD COLUMN offer_days INTEGER",
        "ALTER TABLE bonus_links ADD COLUMN price_variant TEXT",
        "ALTER TABLE bonus_links ADD COLUMN offer_code TEXT",
    ]:
        try:
            con.execute(stmt)
        except Exception:
            pass

    con.commit()
    con.close()

def summarize_buyer(buyer_id: str) -> Dict[str, Any]:
    con = db()
    cur = con.cursor()
    cur.execute("SELECT buyer_name FROM buyers WHERE buyer_id=?", (buyer_id,))
    row = cur.fetchone()
    buyer_name = row[0] if row else None
    cur.execute("SELECT event_type FROM events WHERE buyer_id=?", (buyer_id,))
    events = [r[0] for r in cur.fetchall()]
    con.close()
    purchases = sum(1 for e in events if e=="purchase")
    reviews = sum(1 for e in events if e=="review")
    return {"buyer_name": buyer_name, "purchases": purchases, "reviews": reviews, "segment": "repeat" if purchases>=2 else "new"}

def make_personalized_copy(day: str, buyer_id: str) -> str:
    s = summarize_buyer(buyer_id)
    name = (s.get("buyer_name") or "").strip()
    greet = f"{name}님" if name else "친구야"
    if s["reviews"] >= 2:
        tone = "💎 VIP 감사. 오늘은 ‘나를 특별하게 대하는 행동’ 하나만 해요."
    elif s["reviews"] >= 1:
        tone = "💎 늘 함께해줘서 고마워요. 오늘은 ‘내 편’인 시간을 더 길게 가져도 돼요."
    elif s["purchases"] >= 2:
        tone = "🌿 다시 돌아온 걸 환영해요. 이번 주는 ‘조급함’을 내려놓는 게 포인트예요."
    else:
        tone = "🌱 첫 시작이 제일 소중해요. 오늘은 작은 성공 하나만 만들어봐요."
    return f"{greet} · {day} 보너스\n{tone}"

def weekday_kor(d: date) -> str:
    return "월화수목금토일"[d.weekday()]

def find_reco(tracker: Path, segment: str, platform: str, wday: str) -> Dict[str, str]:
    out = {"mood":"힐링", "color":"민트", "price":"3900", "cta":"즉시 다운로드"}
    if not tracker.exists():
        return out
    try:
        wb = openpyxl.load_workbook(tracker)
        if "Segment_Recommendations" not in wb.sheetnames:
            return out
        ws = wb["Segment_Recommendations"]
        headers = [safe_str(c.value) for c in ws[1]]
        idx = {h:i for i,h in enumerate(headers) if h}
        def get(row, name):
            i = idx.get(name, None)
            return safe_str(row[i]) if i is not None and i < len(row) else ""
        for r in ws.iter_rows(min_row=2, values_only=True):
            if get(r,"segment").lower()==segment.lower() and get(r,"platform").lower()==platform.lower() and get(r,"weekday")==wday:
                return {
                    "mood": get(r,"mood") or out["mood"],
                    "color": get(r,"color") or out["color"],
                    "price": get(r,"price") or out["price"],
                    "cta": get(r,"cta") or out["cta"],
                }
        return out
    except Exception:
        return out


# ---------- Price A/B (3900 vs 4900) – 결과 기반 선택 ----------

def read_price_ab_stats(tracker: Path, segment: str, platform: str, wday: str, season: str="") -> Optional[Dict[str, Any]]:
    """
    Sheet: Price_AB_Stats (v18+)
    Decision metric:
      - PLATFORM_DECISION_METRIC: ev_clickers for SNS, ev_links for store/web
      - then weighted by PLATFORM_EV_WEIGHT × WEEKDAY_EV_WEIGHT × SEASON_EV_WEIGHT
    Optional: season column (if present). If absent, season filter ignored.
    """
    if not tracker.exists():
        return None
    try:
        wb = openpyxl.load_workbook(tracker)
        if "Price_AB_Stats" not in wb.sheetnames:
            return None
        ws = wb["Price_AB_Stats"]
        headers = [safe_str(c.value).lower() for c in ws[1]]
        idxm = {h:i for i,h in enumerate(headers) if h}
        def get(row, name, default=None):
            i = idxm.get(name, None)
            return row[i] if i is not None and i < len(row) else default

        has_season = "season" in idxm
        metric = PLATFORM_DECISION_METRIC.get(platform.lower(), "ev_links")

        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if safe_str(get(r,"segment","")).lower() != segment.lower():
                continue
            if safe_str(get(r,"platform","")).lower() != platform.lower():
                continue
            if safe_str(get(r,"weekday","")) != wday:
                continue
            if has_season and season:
                if safe_str(get(r,"season","")).lower() != season.lower():
                    continue

            variant = safe_str(get(r,"variant","A")).upper() or "A"
            price = int(float(get(r,"price", 3900 if variant=="A" else 4900) or 0))
            ev_links = float(get(r,"ev_links",0.0) or 0.0)
            ev_clickers = float(get(r,"ev_clickers",0.0) or 0.0)
            base = ev_clickers if metric=="ev_clickers" else ev_links
            season_val = season.lower() if season else (safe_str(get(r,"season","")).lower())
            rows.append({
                "variant": variant,
                "price": price,
                "ev_links": ev_links,
                "ev_clickers": ev_clickers,
                "weighted_ev": weighted_ev(platform, wday, season_val, base),
            })
        if not rows:
            return None
        return sorted(rows, key=lambda x: x["weighted_ev"], reverse=True)[0]
    except Exception:
        return None

def get_or_assign_price_variant(buyer_id: str, segment: str, platform: str, wday: str, season: str) -> Tuple[str, int, str]:
    """
    - payload에 price 없을 때 호출.
    - 1) tracker의 Price_AB_Stats가 있으면 EV 최대 variant 선택(하지만 buyer 단위로 고정)
    - 2) 없으면 50/50(요일+buyer_id 해시)로 고정
    returns (variant, price, tone) where tone is 'premium' or 'light'
    """
    # already assigned?
    con = db()
    cur = con.cursor()
    cur.execute("SELECT variant, price FROM ab_price_assign WHERE buyer_id=? AND platform=? AND weekday=? AND segment=?",
                (buyer_id, platform, wday, segment))
    row = cur.fetchone()
    if row:
        con.close()
        v, p = row[0], int(row[1])
        tone = "premium" if p >= 4900 else "light"
        return v, p, tone

    # decide variant
    best = read_price_ab_stats(TRACKER_XLSX, segment, platform, wday, season)
    if best:
        v = best["variant"]
        p = int(best["price"])
    else:
        # deterministic 50/50
        h = sum(ord(c) for c in (buyer_id + platform + wday + segment))
        v = "A" if (h % 2 == 0) else "B"
        p = 3900 if v=="A" else 4900

    cur.execute("INSERT OR REPLACE INTO ab_price_assign(buyer_id,platform,weekday,segment,variant,price,assigned_at) VALUES(?,?,?,?,?,?,?)",
                (buyer_id, platform, wday, segment, v, p, time.time()))

    # Schema evolution (safe ALTER)
    for stmt in [
        "ALTER TABLE bonus_links ADD COLUMN season TEXT",
        "ALTER TABLE bonus_links ADD COLUMN offer_days INTEGER",
        "ALTER TABLE bonus_links ADD COLUMN price_variant TEXT",
        "ALTER TABLE bonus_links ADD COLUMN offer_code TEXT",
    ]:
        try:
            con.execute(stmt)
        except Exception:
            pass

    con.commit()
    con.close()
    tone = "premium" if p >= 4900 else "light"
    return v, p, tone
# ---------- Preset overlay (same as v14) ----------
def pick_font(size: int):
    if FONT_PATH and Path(FONT_PATH).exists():
        return ImageFont.truetype(FONT_PATH, size=size)
    return ImageFont.load_default()

def color_to_rgb(name: str):
    m = {"민트": (140,226,210),"핑크": (255,182,193),"하늘": (164,211,255),"라벤더": (205,180,219),
         "노랑": (255,234,167),"오렌지": (255,198,120),"코랄": (255,160,160),"초록": (166,226,168),
         "블루": (120,170,255),"보라": (175,140,255)}
    return m.get(name.strip(), (200,200,200))

def draw_round_rect(draw, xy, radius, fill):
    draw.rounded_rectangle(list(xy), radius=radius, fill=fill)

def overlay_with_preset(in_png: Path, out_png: Path, main_text: str, preset_name: str, mood: str, color_name: str, price: str, cta: str):
    im = Image.open(in_png).convert("RGBA")
    W,H = im.size
    draw = ImageDraw.Draw(im)
    preset = PRESETS.get(preset_name, PRESETS.get("top", {}))
    box_w = int(W * float(preset.get("box_width_ratio", 0.86)))
    x = (W - box_w) // 2
    y = int(H * float(preset.get("box_y_ratio", 0.08)))
    pad = int(preset.get("padding", 26))
    font_size = int(preset.get("font_size", 44))
    line_h = int(preset.get("line_height", 56))
    box_rgba = tuple(preset.get("box_rgba", [255,255,255,185]))
    text_rgba = tuple(preset.get("text_rgba", [40,35,32,255]))
    font = pick_font(font_size)

    # wrap
    lines = []
    for raw in (main_text or "").split("\n"):
        raw = raw.strip()
        if not raw:
            lines.append("")
            continue
        cur = ""
        for ch in raw:
            test = cur + ch
            if draw.textlength(test, font=font) <= (box_w - pad*2):
                cur = test
            else:
                lines.append(cur)
                cur = ch
        if cur: lines.append(cur)
    if len(lines) < 2: lines.append("")
    box_h = pad*2 + line_h*len(lines)

    box = Image.new("RGBA", (box_w, box_h), box_rgba)
    im.alpha_composite(box, (x,y))

    tx, ty = x + pad, y + pad
    for ln in lines:
        draw.text((tx,ty), ln, font=font, fill=text_rgba)
        ty += line_h

    # badges
    badge_cfg = preset.get("badge", {"enabled": True})
    if badge_cfg.get("enabled", True):
        by = int(H * float(badge_cfg.get("y_ratio", 0.02)))
        bx = x
        badge_h = int(font_size * 1.25)
        badge_pad_x = 18
        badge_gap = 12

        def badge(label: str):
            nonlocal bx
            f = pick_font(int(font_size*0.72))
            w = int(draw.textlength(label, font=f)) + badge_pad_x*2
            draw_round_rect(draw, (bx, by, bx+w, by+badge_h), radius=badge_h//2, fill=(255,255,255,210))
            draw.text((bx+badge_pad_x, by+int(badge_h*0.22)), label, font=f, fill=(30,30,30,255))
            bx += w + badge_gap

        badge(f"mood · {mood}")
        badge(f"color · {color_name}")
        c = color_to_rgb(color_name)
        sw = badge_h
        draw_round_rect(draw, (bx, by, bx+sw, by+sw), radius=sw//3, fill=(c[0],c[1],c[2],255))
        bx += sw + badge_gap
        badge(f"{price}원")
        badge(f"{cta}")

    out_png.parent.mkdir(parents=True, exist_ok=True)
    im.convert("RGB").save(out_png, "PNG")

# ---------- XLSX override (핵심) ----------
def override_cards_xlsx(src_xlsx: Path, out_xlsx: Path, sheet: str, day: str, mood: str, color: str, price: str, cta: str):
    wb = openpyxl.load_workbook(src_xlsx)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    headers = [safe_str(c.value).lower() for c in ws[1]]
    col = {h:i+1 for i,h in enumerate(headers) if h}
    # ensure optional columns exist
    def ensure_col(name: str):
        if name not in col:
            ws.cell(row=1, column=ws.max_column+1, value=name)
            col[name] = ws.max_column
    for k in ["mood","color","price","cta"]:
        ensure_col(k)
    day_col = col.get("day", None)
    if not day_col:
        raise ValueError("Cards sheet must have 'day' header")
    target_row = None
    for r in range(2, ws.max_row+1):
        v = safe_str(ws.cell(row=r, column=day_col).value).upper()
        if v.startswith("DAY"):
            v = v
        if v == day:
            target_row = r
            break
        # accept 9 or 09
        digits = "".join(ch for ch in v if ch.isdigit())
        if digits and f"DAY{int(digits):02d}" == day:
            target_row = r
            break
    if target_row is None:
        # append new row
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=day_col, value=day)
    ws.cell(row=target_row, column=col["mood"], value=mood)
    ws.cell(row=target_row, column=col["color"], value=color)
    ws.cell(row=target_row, column=col["price"], value=price)
    ws.cell(row=target_row, column=col["cta"], value=cta)
    wb.save(out_xlsx)

# ---------- generation via patched run_generate.py ----------
def locate_day_png(out_dir: Path, day: str) -> Optional[Path]:
    m = sorted(out_dir.rglob(f"{day}*.png"))
    return m[0] if m else None

def generate_bonus_day(day: str, platform: str, override_xlsx: Path) -> Path:
    if not RUN_GENERATE.exists():
        raise RuntimeError(f"run_generate.py not found at {RUN_GENERATE}")
    out_dir = BONUS_OUT_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{platform}"
    cmd = [
        sys.executable, str(RUN_GENERATE),
        "--season", BONUS_SEASON,
        "--platform", platform,
        "--format", "reels" if platform=="instagram" else "shorts",
        "--mode", "paid",
        "--days", str(int(day.replace("DAY",""))),
        "--xlsx", str(override_xlsx),
        "--sheet", "Cards",
        "--thumb_pick", "A",
        "--base_url", "https://example.com/buyer",
        "--utm_campaign", "bonus_gen",
        "--font_path", FONT_PATH or "",
        "--out_dir", str(out_dir),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f"run_generate failed:\nSTDOUT:\n{r.stdout}\nSTDERR:\n{r.stderr}")
    png = locate_day_png(out_dir, day)
    if not png:
        raise RuntimeError(f"Could not locate generated image for {day} under {out_dir}")
    return png

# ---------- Upload adapters ----------
def upload_to_s3(local_path: Path) -> Optional[str]:
    bucket = os.environ.get("S3_BUCKET","").strip()
    if not bucket:
        return None
    try:
        import boto3
        region = os.environ.get("AWS_REGION","ap-northeast-2")
        s3 = boto3.client("s3", region_name=region)
        key = f"alloki-dalloki/bonus/{local_path.name}"
        s3.upload_file(str(local_path), bucket, key, ExtraArgs={"ContentType":"image/png"})
        return f"https://{bucket}.s3.{region}.amazonaws.com/{key}"
    except Exception as e:
        print("S3 upload failed:", e, file=sys.stderr)
        return None

def upload_adapter(local_path: Path) -> str:
    url = upload_to_s3(local_path)
    if url: return url
    return f"https://cdn.example.com/{local_path.name}"

# ---------- Tracking + tracker writeback ----------
def issue_tracking_link(day: str, buyer_id: str, platform: str, target_url: str, base_url: str) -> str:
    token = secrets.token_urlsafe(12)
    con = db()
    con.execute("INSERT OR REPLACE INTO bonus_links(token,buyer_id,day,target_url,platform,created_at,clicks) VALUES(?,?,?,?,?,?,0)",
                (token, buyer_id, day, target_url, platform, time.time()))

    # Schema evolution (safe ALTER)
    for stmt in [
        "ALTER TABLE bonus_links ADD COLUMN season TEXT",
        "ALTER TABLE bonus_links ADD COLUMN offer_days INTEGER",
        "ALTER TABLE bonus_links ADD COLUMN price_variant TEXT",
        "ALTER TABLE bonus_links ADD COLUMN offer_code TEXT",
    ]:
        try:
            con.execute(stmt)
        except Exception:
            pass

    con.commit()
    con.close()
    return f"{base_url}/r/{day.lower()}/{token}"

def ensure_bonus_sheet(wb):
    name = "Bonus_Clicks"
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(["timestamp","day","token","buyer_id","platform","clicks","target_url"])
        return ws
    return wb[name]

def append_bonus_click(tracker_path: Path, day: str, token: str, buyer_id: str, platform: str, clicks: int, target_url: str):
    if not tracker_path.exists():
        return
    wb = openpyxl.load_workbook(tracker_path)
    ws = ensure_bonus_sheet(wb)
    ws.append([datetime.now().isoformat(timespec="seconds"), day, token, buyer_id, platform, clicks, target_url])
    wb.save(tracker_path)


# ---------- Monthly Price_AB_Stats auto update ----------
def month_range_utc(year: int, month: int) -> Tuple[float, float]:
    """
    Returns (start_ts, end_ts) in epoch seconds for the given month in local TIMEZONE,
    but stored as epoch seconds (server uses epoch seconds everywhere).
    Note: This is "good enough" if server runs in same TZ; for strict TZ handling use zoneinfo.
    """
    # Use naive local time; in Korea it's fine (no DST). If you need DST, switch to zoneinfo.
    start = datetime(year, month, 1, 0, 0, 0)
    if month == 12:
        end = datetime(year+1, 1, 1, 0, 0, 0)
    else:
        end = datetime(year, month+1, 1, 0, 0, 0)
    return start.timestamp(), end.timestamp()

def prev_month_year_month(now: datetime) -> Tuple[int, int]:
    if now.month == 1:
        return now.year - 1, 12
    return now.year, now.month - 1

def ensure_price_ab_sheet(wb):
    name = "Price_AB_Stats"
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(["segment","platform","weekday","variant","price","month","links_issued","clicks","unique_clickers","click_rate","conversions_total","conv_rate_links","click_cvr","conv_purchase","conv_coupon","conv_revisit","ev_links","ev_clickers"])
        return ws
    return wb[name]

def update_price_ab_stats_for_month(tracker_path: Path, year: int, month: int):
    """
    완전 무인: DB(clicks/bonus_links/events/ab_price_assign) 기반으로
    - links_issued: 해당 월에 발급된 bonus 링크 수(DAY09/DAY10)
    - clicks: 해당 월 클릭 수
    - click_rate: clicks / links_issued
    - conversions: 클릭 후 CONV_WINDOW_DAYS 내 "추가 구매"(repeat purchase) 발생 수
    - conv_rate: conversions / links_issued  (구매율)
    그리고 Price_AB_Stats 시트에 month(YYYY-MM) 단위로 overwrite
    """
    if not tracker_path.exists():
        return

    start_ts, end_ts = month_range_utc(year, month)
    month_key = f"{year:04d}-{month:02d}"

    con = db()
    cur = con.cursor()

    # 1) 이번 달 배정된 A/B 그룹 (exposure group)
    cur.execute("""
        SELECT buyer_id, platform, weekday, segment, variant, price, assigned_at
        FROM ab_price_assign
        WHERE assigned_at >= ? AND assigned_at < ?
    """, (start_ts, end_ts))
    assigns = cur.fetchall()

    # group dict
    groups = {}  # (segment, platform, weekday, variant, price) -> set(buyers)
    for buyer_id, platform, weekday, segment, variant, price, assigned_at in assigns:
        key = (segment, platform, weekday, variant, int(price))
        groups.setdefault(key, set()).add(buyer_id)

    # Helper: count links issued and clicks for buyers within month
    def links_and_clicks(buyers: set, platform: str) -> Tuple[int, int, int, Dict[str, float]]:
        if not buyers:
            return 0, 0, 0, {}
        buyers_list = tuple(buyers)
        # links issued (DAY09/DAY10 only) in this month
        q_links = f"""
            SELECT token, buyer_id
            FROM bonus_links
            WHERE created_at >= ? AND created_at < ?
              AND platform = ?
              AND day IN ('DAY09','DAY10')
              AND buyer_id IN ({",".join(["?"]*len(buyers_list))})
        """
        cur.execute(q_links, (start_ts, end_ts, platform, *buyers_list))
        link_rows = cur.fetchall()
        tokens = [r[0] for r in link_rows]
        links_issued = len(tokens)
        if not tokens:
            return links_issued, 0, 0, {}

        # clicks in this month for those tokens
        tokens_list = tuple(tokens)
        q_clicks = f"""
            SELECT token, buyer_id, MIN(ts) as first_click_ts, COUNT(*) as c
            FROM clicks
            WHERE ts >= ? AND ts < ?
              AND token IN ({",".join(["?"]*len(tokens_list))})
            GROUP BY token, buyer_id
        """
        cur.execute(q_clicks, (start_ts, end_ts, *tokens_list))
        click_rows = cur.fetchall()
        clicks = sum(int(r[3]) for r in click_rows)
        # first click per buyer (min across tokens)
        first_click_by_buyer = {}
        for token, buyer_id, first_ts, c in click_rows:
            if buyer_id not in first_click_by_buyer:
                first_click_by_buyer[buyer_id] = float(first_ts)
            else:
                first_click_by_buyer[buyer_id] = min(first_click_by_buyer[buyer_id], float(first_ts))
        unique_clickers = len(first_click_by_buyer)
        return links_issued, clicks, unique_clickers, first_click_by_buyer

    # 2) conversions: "추가 구매" = click 이후 CONV_WINDOW_DAYS 내 purchase event 존재 (해당 월 배정 group 안에서)
    rows_out = []
    for (segment, platform, weekday, variant, price), buyers in groups.items():
        links_issued, clicks, unique_clickers, first_click_map = links_and_clicks(buyers, platform)
        click_rate = (clicks / links_issued) if links_issued else 0.0

        conv_purchase = 0
        conv_coupon = 0
        conv_revisit = 0
        window_sec = CONV_WINDOW_DAYS * 86400

        # conversion = click 이후 window 내 (purchase OR coupon OR revisit) 중 하나라도 발생하면 conversion 인정
        for buyer_id, first_click_ts in first_click_map.items():
            # purchase
            cur.execute("""
                SELECT COUNT(*)
                FROM events
                WHERE buyer_id=? AND event_type='purchase'
                  AND created_at > ? AND created_at <= ?
            """, (buyer_id, first_click_ts, first_click_ts + window_sec))
            p = int(cur.fetchone()[0] or 0)
            if p > 0:
                conv_purchase += 1

            # coupon (쿠폰 사용 이벤트)
            cur.execute("""
                SELECT COUNT(*)
                FROM events
                WHERE buyer_id=? AND event_type IN ('coupon','coupon_redeem','redeem','coupon_use')
                  AND created_at > ? AND created_at <= ?
            """, (buyer_id, first_click_ts, first_click_ts + window_sec))
            c = int(cur.fetchone()[0] or 0)
            if c > 0:
                conv_coupon += 1

            # revisit (재방문 이벤트)
            cur.execute("""
                SELECT COUNT(*)
                FROM events
                WHERE buyer_id=? AND event_type IN ('revisit','return','visit','pageview')
                  AND created_at > ? AND created_at <= ?
            """, (buyer_id, first_click_ts, first_click_ts + window_sec))
            r = int(cur.fetchone()[0] or 0)
            if r > 0:
                conv_revisit += 1

        # total conversions: 구매/쿠폰/재방문 중 1개라도 해당되면 conversion으로 카운트(중복 제거)
        conversions_total = 0
        for buyer_id, first_click_ts in first_click_map.items():
            cur.execute("""
                SELECT COUNT(*)
                FROM events
                WHERE buyer_id=?
                  AND event_type IN ('purchase','coupon','coupon_redeem','redeem','coupon_use','revisit','return','visit','pageview')
                  AND created_at > ? AND created_at <= ?
            """, (buyer_id, first_click_ts, first_click_ts + window_sec))
            anycnt = int(cur.fetchone()[0] or 0)
            if anycnt > 0:
                conversions_total += 1

        conv_rate_links = (conversions_total / links_issued) if links_issued else 0.0
        click_cvr = (conversions_total / unique_clickers) if unique_clickers else 0.0

        ev_links = float(price) * conv_rate_links
        ev_clickers = float(price) * click_cvr

        rows_out.append({
            "segment": segment,
            "platform": platform,
            "weekday": weekday,
            "variant": variant,
            "price": int(price),
            "month": month_key,
            "links_issued": links_issued,
            "clicks": clicks,
            "unique_clickers": unique_clickers,
            "click_rate": round(click_rate, 6),
            "conversions_total": conversions_total,
            "conv_rate_links": round(conv_rate_links, 6),
            "click_cvr": round(click_cvr, 6),
            "conv_purchase": conv_purchase,
            "conv_coupon": conv_coupon,
            "conv_revisit": conv_revisit,
            "ev_links": round(ev_links, 6),
            "ev_clickers": round(ev_clickers, 6),
        })

    con.close()

    # 3) write back to tracker_xlsx (overwrite rows for this month+keys)
    wb = openpyxl.load_workbook(tracker_path)
    ws = ensure_price_ab_sheet(wb)

    # build header map
    headers = [safe_str(c.value) for c in ws[1]]
    col = {h:i+1 for i,h in enumerate(headers) if h}

    # index existing rows to delete (same month and same key)
    to_delete = []
    for r in range(2, ws.max_row+1):
        seg = safe_str(ws.cell(r, col.get("segment",1)).value)
        plat = safe_str(ws.cell(r, col.get("platform",2)).value)
        wd = safe_str(ws.cell(r, col.get("weekday",3)).value)
        var = safe_str(ws.cell(r, col.get("variant",4)).value)
        pr  = safe_str(ws.cell(r, col.get("price",5)).value)
        mth = safe_str(ws.cell(r, col.get("month",6)).value)
        if mth == month_key:
            to_delete.append(r)
    # delete bottom-up
    for r in reversed(to_delete):
        ws.delete_rows(r, 1)

    # append new rows
    for item in rows_out:
        ws.append([
            item["segment"], item["platform"], item["weekday"], item["variant"], item["price"], item["month"],
            item["links_issued"], item["clicks"], item["unique_clickers"], item["click_rate"],
            item["conversions_total"], item["conv_rate_links"], item["click_cvr"],
            item["conv_purchase"], item["conv_coupon"], item["conv_revisit"],
            item["ev_links"], item["ev_clickers"]
        ])

    wb.save(tracker_path)


def ensure_offer_sheet(wb):
    name = "Offer_Stats"
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(["segment","platform","weekday","season","offer_code","offer_days","price","month","links_issued","clicks","unique_clickers","conversions_total","conv_rate_links","click_cvr","ev_links","ev_clickers"])
        return ws
    return wb[name]

def parse_offer_days_from_product(product_name: str) -> int:
    s = safe_str(product_name)
    for n in (7,14,21):
        if f"{n}일" in s or f"{n}-day" in s or f"{n}day" in s:
            return n
    return 0

def update_offer_stats_for_month(tracker_path: Path, year: int, month: int):
    if not tracker_path.exists():
        return
    start_ts, end_ts = month_range_utc(year, month)
    month_key = f"{year:04d}-{month:02d}"
    con = db()
    cur = con.cursor()

    cur.execute("""
        SELECT buyer_id, platform, season, offer_code, offer_days, created_at
        FROM bonus_links
        WHERE created_at >= ? AND created_at < ?
          AND day IN ('DAY09','DAY10')
    """, (start_ts, end_ts))
    bl = cur.fetchall()

    def weekday_from_ts(ts: float) -> str:
        d = datetime.fromtimestamp(ts)
        return "월화수목금토일"[d.weekday()]

    groups = {}
    for buyer_id, platform, season, offer_code, offer_days, created_at in bl:
        seg = summarize_buyer(buyer_id)["segment"]
        wd = weekday_from_ts(float(created_at))
        oc = safe_str(offer_code).upper() or ("D21" if int(offer_days or 0)==21 else ("D14" if int(offer_days or 0)==14 else ("D7" if int(offer_days or 0)==7 else "")))
        if not oc:
            oc = "SEASONPACK"
        key = (seg, safe_str(platform).lower(), wd, safe_str(season).lower(), oc, int(offer_days or 0))
        groups.setdefault(key, set()).add(buyer_id)

    window_sec = CONV_WINDOW_DAYS * 86400
    rows_out = []

    offer_price_map = OFFER_PRICE_MAP

    for (seg, platform, wd, season, offer_code, offer_days), buyers in groups.items():
        if not buyers:
            continue
        buyers_list = tuple(buyers)
        q_links = f"""
            SELECT token, buyer_id
            FROM bonus_links
            WHERE created_at >= ? AND created_at < ?
              AND platform = ?
              AND day IN ('DAY09','DAY10')
              AND offer_days = ?
              AND (offer_code = ? OR offer_code IS NULL OR offer_code='')
              AND buyer_id IN ({",".join(["?"]*len(buyers_list))})
        """
        cur.execute(q_links, (start_ts, end_ts, platform, offer_days, offer_code, *buyers_list))
        link_rows = cur.fetchall()
        tokens = [r[0] for r in link_rows]
        links_issued = len(tokens)
        if not tokens:
            continue

        tokens_list = tuple(tokens)
        q_clicks = f"""
            SELECT token, buyer_id, MIN(ts) as first_click_ts, COUNT(*) as c
            FROM clicks
            WHERE ts >= ? AND ts < ?
              AND token IN ({",".join(["?"]*len(tokens_list))})
            GROUP BY token, buyer_id
        """
        cur.execute(q_clicks, (start_ts, end_ts, *tokens_list))
        click_rows = cur.fetchall()
        clicks = sum(int(r[3]) for r in click_rows)
        first_click_by_buyer = {}
        for token, buyer_id, first_ts, c in click_rows:
            if buyer_id not in first_click_by_buyer:
                first_click_by_buyer[buyer_id] = float(first_ts)
            else:
                first_click_by_buyer[buyer_id] = min(first_click_by_buyer[buyer_id], float(first_ts))
        unique_clickers = len(first_click_by_buyer)

        conversions_total = 0
        for buyer_id, first_click_ts in first_click_by_buyer.items():
            cur.execute("""
                SELECT event_type, product_name
                FROM events
                WHERE buyer_id=?
                  AND created_at > ? AND created_at <= ?
                  AND event_type IN ('purchase','coupon','coupon_redeem','redeem','coupon_use','revisit','return','visit','pageview')
            """, (buyer_id, first_click_ts, first_click_ts + window_sec))
            evs = cur.fetchall()
            ok = False
            for et, pn in evs:
                et = safe_str(et).lower()
                if et == "purchase":
                    if offer_code == 'SEASONPACK':
                        if ('시즌' in safe_str(pn)) or ('season' in safe_str(pn).lower()):
                            ok = True
                            break
                    else:
                        if offer_days == 0 or parse_offer_days_from_product(pn) == offer_days:
                        ok = True
                        break
                else:
                    ok = True
                    break
            if ok:
                conversions_total += 1

        conv_rate_links = conversions_total / links_issued if links_issued else 0.0
        click_cvr = conversions_total / unique_clickers if unique_clickers else 0.0

        price = int(offer_price_map.get(offer_code, 0))
        ev_links = price * conv_rate_links
        ev_clickers = price * click_cvr

        rows_out.append([seg, platform, wd, season, offer_days, price, month_key, links_issued, clicks, unique_clickers, conversions_total,
                         round(conv_rate_links,6), round(click_cvr,6), round(ev_links,6), round(ev_clickers,6)])

    con.close()

    wb = openpyxl.load_workbook(tracker_path)
    ws = ensure_offer_sheet(wb)
    headers = [safe_str(c.value).lower() for c in ws[1]]
    idxm = {h:i+1 for i,h in enumerate(headers) if h}
    mcol = idxm.get("month", 7)

    del_rows = []
    for r in range(2, ws.max_row+1):
        if safe_str(ws.cell(r, mcol).value) == month_key:
            del_rows.append(r)
    for r in reversed(del_rows):
        ws.delete_rows(r,1)

    for row in rows_out:
        ws.append(row)

    wb.save(tracker_path)


def monthly_worker_loop():
    # Run once on startup for previous month (optional), then sleep-check every hour.
    try:
        now = datetime.now()
        y, m = prev_month_year_month(now)
        update_price_ab_stats_for_month(TRACKER_XLSX, y, m)
        update_offer_stats_for_month(TRACKER_XLSX, y, m)
    except Exception as e:
        print("monthly stats initial run failed:", e, file=sys.stderr)

    last_ran_month = None
    while True:
        try:
            now = datetime.now()
            # On the 1st day at configured time, update previous month
            if now.day == 1 and now.hour == MONTHLY_STATS_HOUR and now.minute >= MONTHLY_STATS_MINUTE:
                y, m = prev_month_year_month(now)
                key = f"{y:04d}-{m:02d}"
                if key != last_ran_month:
                    update_price_ab_stats_for_month(TRACKER_XLSX, y, m)
        update_offer_stats_for_month(TRACKER_XLSX, y, m)
                    last_ran_month = key
        except Exception as e:
            print("monthly stats loop failed:", e, file=sys.stderr)
        time.sleep(3600)

# ---------- Webhooks ----------

@APP.post("/webhook/event")
async def webhook_event(req: Request):
    """
    Generic event logger:
    payload: {
      "buyer_id": "...",
      "platform": "instagram|tiktok",
      "event_type": "coupon|revisit|purchase|review|...",
      "order_id": "... optional",
      "product_name": "... optional",
      "buyer_name": "... optional"
    }
    """
    payload = await req.json()
    buyer_id = safe_str(payload.get("buyer_id",""))
    if not buyer_id:
        return JSONResponse({"ok": False, "error": "buyer_id required"}, status_code=400)
    platform = (safe_str(payload.get("platform","instagram")) or "instagram").lower()
    event_type = safe_str(payload.get("event_type","")).lower()
    if not event_type:
        return JSONResponse({"ok": False, "error": "event_type required"}, status_code=400)
    buyer_name = safe_str(payload.get("buyer_name","")) or None
    order_id = safe_str(payload.get("order_id",""))
    product_name = safe_str(payload.get("product_name","알록이 달록이 카드"))

    con = db()
    con.execute("INSERT OR IGNORE INTO buyers(buyer_id,buyer_name,created_at) VALUES(?,?,?)", (buyer_id, buyer_name, time.time()))
    if buyer_name:
        con.execute("UPDATE buyers SET buyer_name=? WHERE buyer_id=?", (buyer_name, buyer_id))
    con.execute("INSERT INTO events(buyer_id,event_type,platform,order_id,product_name,created_at) VALUES(?,?,?,?,?,?)",
                (buyer_id, event_type, platform, order_id, product_name, time.time()))
    con.commit(); con.close()
    return JSONResponse({"ok": True})

@APP.post("/webhook/purchase")
async def webhook_purchase(req: Request):
    payload = await req.json()
    buyer_id = safe_str(payload.get("buyer_id","")) or f"buyer_{int(time.time())}"
    buyer_name = safe_str(payload.get("buyer_name","")) or None
    platform = (safe_str(payload.get("platform","instagram")) or "instagram").lower()

    con = db()
    con.execute("INSERT OR IGNORE INTO buyers(buyer_id,buyer_name,created_at) VALUES(?,?,?)", (buyer_id, buyer_name, time.time()))
    if buyer_name:
        con.execute("UPDATE buyers SET buyer_name=? WHERE buyer_id=?", (buyer_name, buyer_id))
    con.execute("INSERT INTO events(buyer_id,event_type,platform,order_id,product_name,created_at) VALUES(?,?,?,?,?,?)",
                (buyer_id, "purchase", platform, safe_str(payload.get("order_id","")), safe_str(payload.get("product_name","알록이 달록이 카드")), time.time()))
    con.commit(); con.close()

    seg = summarize_buyer(buyer_id)["segment"]
    wday = weekday_kor(date.today())
    reco = find_reco(TRACKER_XLSX, seg, platform, wday)

    # payload 우선
    mood = safe_str(payload.get("mood","")) or reco["mood"]
    color = safe_str(payload.get("color","")) or reco["color"]
    cta  = safe_str(payload.get("cta",""))  or reco["cta"]

    price_payload = safe_str(payload.get("price",""))
    if price_payload:
        price = price_payload
        price_variant = "MANUAL"
        price_tone = "premium" if int("".join(ch for ch in price if ch.isdigit()) or 0) >= 4900 else "light"
    else:
        v, p, t = get_or_assign_price_variant(buyer_id, seg, platform, wday, season)
        price_variant, price, price_tone = v, str(p), t

    preset = safe_str(payload.get("preset",""))
    offer_code = safe_str(payload.get("offer_code","")) or ""
    offer_days_in = int(payload.get("offer_days") or 0)
    if offer_code:
        offer_days = offer_code_to_days(offer_code)
    elif offer_days_in:
        offer_days = offer_days_in
        offer_code = "D21" if offer_days==21 else ("D14" if offer_days==14 else "D7")
    else:
        offer_code, offer_days = choose_offer(buyer_id, seg, platform, wday, season) or (DEFAULT_PRESET_INSTAGRAM if platform=="instagram" else DEFAULT_PRESET_TIKTOK)
    if preset not in PRESETS: preset = "top"

    # --- 핵심: 임시 xlsx overwrite로 prompt까지 반영 ---
    tmp = BONUS_OUT_DIR / f"tmp_cards_{buyer_id}_{int(time.time())}.xlsx"
    override_cards_xlsx(CARDS_XLSX, tmp, "Cards", "DAY09", mood, color, price, cta)

    raw = generate_bonus_day("DAY09", platform, tmp)
    main_text = make_personalized_copy("DAY09", buyer_id)

    out_png = BONUS_OUT_DIR / f"DAY09_{buyer_id}_{int(time.time())}.png"
    overlay_with_preset(raw, out_png, main_text, preset, mood, color, price, cta)

    target_url = upload_adapter(out_png)
    base_url = str(req.base_url).rstrip("/")
    track = issue_tracking_link("DAY09", buyer_id, platform, target_url, base_url, season=season, offer_days=offer_days, price_variant=price_variant, offer_code=offer_code)

    return JSONResponse({
        "ok": True,
        "segment": seg,
        "personalization": {"season": season, "offer_code": offer_code, "offer_days": offer_days, "mood": mood, "color": color, "price": price, "price_variant": price_variant, "price_tone": price_tone, "cta": cta, "preset": preset},
        "day09_tracking_link": track,
        "coupon": f"ALLD-10-{secrets.token_hex(2).upper()}",
        "note": "요일·플랫폼 최적 mood/color 자동 + price A/B 분기 + 프롬프트 톤 분기(v16)"
    })

@APP.post("/webhook/review")
async def webhook_review(req: Request):
    payload = await req.json()
    buyer_id = safe_str(payload.get("buyer_id",""))
    if not buyer_id:
        return JSONResponse({"ok": False, "error": "buyer_id required"}, status_code=400)
    platform = (safe_str(payload.get("platform","instagram")) or "instagram").lower()

    con = db()
    con.execute("INSERT OR IGNORE INTO buyers(buyer_id,buyer_name,created_at) VALUES(?,?,?)", (buyer_id, None, time.time()))
    con.execute("INSERT INTO events(buyer_id,event_type,platform,order_id,product_name,created_at) VALUES(?,?,?,?,?,?)",
                (buyer_id, "review", platform, safe_str(payload.get("order_id","")), safe_str(payload.get("product_name","알록이 달록이 카드")), time.time()))
    con.commit(); con.close()

    seg = summarize_buyer(buyer_id)["segment"]
    wday = weekday_kor(date.today())
    reco = find_reco(TRACKER_XLSX, seg, platform, wday)

    mood = safe_str(payload.get("mood","")) or reco["mood"]
    color = safe_str(payload.get("color","")) or reco["color"]
    cta  = safe_str(payload.get("cta",""))  or reco["cta"]

    price_payload = safe_str(payload.get("price",""))
    if price_payload:
        price = price_payload
        price_variant = "MANUAL"
        price_tone = "premium" if int("".join(ch for ch in price if ch.isdigit()) or 0) >= 4900 else "light"
    else:
        v, p, t = get_or_assign_price_variant(buyer_id, seg, platform, wday, season)
        price_variant, price, price_tone = v, str(p), t

    preset = safe_str(payload.get("preset",""))
    offer_code = safe_str(payload.get("offer_code","")) or ""
    offer_days_in = int(payload.get("offer_days") or 0)
    if offer_code:
        offer_days = offer_code_to_days(offer_code)
    elif offer_days_in:
        offer_days = offer_days_in
        offer_code = "D21" if offer_days==21 else ("D14" if offer_days==14 else "D7")
    else:
        offer_code, offer_days = choose_offer(buyer_id, seg, platform, wday, season) or (DEFAULT_PRESET_INSTAGRAM if platform=="instagram" else DEFAULT_PRESET_TIKTOK)
    if preset not in PRESETS: preset = "middle"

    tmp = BONUS_OUT_DIR / f"tmp_cards_{buyer_id}_{int(time.time())}.xlsx"
    override_cards_xlsx(CARDS_XLSX, tmp, "Cards", "DAY10", mood, color, price, cta)

    raw = generate_bonus_day("DAY10", platform, tmp)
    main_text = make_personalized_copy("DAY10", buyer_id)

    out_png = BONUS_OUT_DIR / f"DAY10_{buyer_id}_{int(time.time())}.png"
    overlay_with_preset(raw, out_png, main_text, preset, mood, color, price, cta)

    target_url = upload_adapter(out_png)
    base_url = str(req.base_url).rstrip("/")
    track = issue_tracking_link("DAY10", buyer_id, platform, target_url, base_url, season=season, offer_days=offer_days, price_variant=price_variant, offer_code=offer_code)

    return JSONResponse({
        "ok": True,
        "segment": seg,
        "personalization": {"season": season, "offer_code": offer_code, "offer_days": offer_days, "mood": mood, "color": color, "price": price, "price_variant": price_variant, "price_tone": price_tone, "cta": cta, "preset": preset},
        "day10_tracking_link": track,
        "note": "요일·플랫폼 최적 mood/color 자동 + price A/B 분기 + 프롬프트 톤 분기(v16)"
    })

@APP.get("/r/{day}/{token}")
async def redirect_day(day: str, token: str, req: Request):
    day_norm = day.upper()
    if not day_norm.startswith("DAY"):
        day_norm = "DAY" + day_norm.replace("DAY","").zfill(2)

    con = db()
    cur = con.cursor()
    cur.execute("SELECT buyer_id, target_url, platform, clicks FROM bonus_links WHERE token=?", (token,))
    row = cur.fetchone()
    if not row:
        con.close()
        return JSONResponse({"ok": False, "error": "invalid token"}, status_code=404)

    buyer_id, target_url, platform, clicks = row
    ua = req.headers.get("user-agent","")
    ref = req.headers.get("referer","")
    cur.execute("UPDATE bonus_links SET clicks=? WHERE token=?", (int(clicks)+1, token))
    cur.execute("INSERT INTO clicks(token,buyer_id,day,platform,ts,ua,ref) VALUES(?,?,?,?,?,?,?)",
                (token, buyer_id, day_norm, platform, time.time(), ua, ref))
    con.commit()
    cur.execute("SELECT clicks FROM bonus_links WHERE token=?", (token,))
    clicks2 = cur.fetchone()[0]
    con.close()

    append_bonus_click(TRACKER_XLSX, day_norm, token, buyer_id, platform, clicks2, target_url)
    return RedirectResponse(target_url, status_code=302)

def main():
    init_db()
    if AUTO_MONTHLY_STATS:
        t = threading.Thread(target=monthly_worker_loop, daemon=True)
        t.start()
    ap = argparse.ArgumentParser()
    ap.add_argument("--host", default="0.0.0.0")
    ap.add_argument("--port", type=int, default=8787)
    args = ap.parse_args()
    uvicorn.run(APP, host=args.host, port=args.port)

if __name__ == "__main__":
    main()
